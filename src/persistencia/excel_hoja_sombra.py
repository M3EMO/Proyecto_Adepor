"""
Hoja "Sombra" — auditoria comparativa Opcion 1 (activa) vs Opcion 4 (shadow).

Un row por partido donde al menos una de las dos opciones registro pick real.
Muestra stake, resultado y P/L de cada opcion, la diferencia y stats globales
al pie para determinar cual estrategia lleva ventaja.

Extraido del motor_sincronizador.py monolitico en fase 4 (2026-04-21).
"""
from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.comun.resolucion import determinar_resultado_entero
from src.persistencia.excel_estilos import fill
from src.persistencia.excel_formulas import cuota_1x2


def crear_hoja_sombra(wb, datos, bankroll):
    """Auditoria comparativa Op1 (activa) vs Op4 (shadow)."""
    ws = wb.create_sheet("Sombra")

    FONT_TITLE = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    FONT_HDR   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    FONT_KPI   = Font(name='Arial', bold=True, size=10)
    FONT_D     = Font(name='Arial', size=10)
    FONT_SUB   = Font(name='Arial', italic=True, size=9, color='595959')

    FILL_HDR1  = fill('2E75B6')
    FILL_HDR4  = fill('548235')
    FILL_NEUTRO_ROW = fill('F5F5F5')
    FILL_BLANCO     = fill('FFFFFF')
    FILL_GANADA_CELL  = fill('C6EFCE')
    FILL_PERDIDA_CELL = fill('FFC7CE')
    FILL_PEND_CELL    = fill('FFEB9C')
    FILL_MEJOR = fill('C6EFCE')
    FILL_PEOR  = fill('FFC7CE')
    FILL_IGUAL = fill('F2F2F2')

    BORDER = Border(
        left=Side(style='thin',  color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin',   color='D9D9D9'),
        bottom=Side(style='thin',color='D9D9D9'),
    )

    # Anchos columnas A..N
    for ci, w in enumerate([13, 30, 12, 20, 11, 12, 13, 20, 11, 12, 13, 13, 9, 9], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # --- Titulo ---
    ws.merge_cells('A1:N1')
    c = ws.cell(1, 1, "AUDITORIA COMPARATIVA: OPCION 1 (ACTIVA) vs OPCION 4 (SHADOW)")
    c.font = FONT_TITLE; c.fill = fill('1F4E79')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:N2')
    c = ws.cell(2, 1,
                "Op1 = Floor 33% + EV escalado + Caminos 2B/3 + Bloqueo empates + xG Margen O/U  (apuesta real)   |   "
                "Op4 = Floor 33% sin EV escalado + fallback prob baja (shadow, solo auditoria)")
    c.font = FONT_SUB; c.fill = fill('D6E4F0')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 14

    # --- Headers fila 3 (bloques Op1/Op4/Diferencia) ---
    hdrs = [
        ('A3:C3', 'PARTIDO',            fill('4472C4')),
        ('D3:G3', 'OPCION 1 (ACTIVA)',  FILL_HDR1),
        ('H3:K3', 'OPCION 4 (SHADOW)',  FILL_HDR4),
        ('L3:N3', 'DIFERENCIA',         fill('7030A0')),
    ]
    for rng, txt, fill_ in hdrs:
        ws.merge_cells(rng)
        c = ws.cell(3, ord(rng[0]) - 64, txt)
        c.font = FONT_HDR; c.fill = fill_
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[3].height = 16

    # --- Headers fila 4 (columnas individuales) ---
    col_hdrs = [
        'Fecha', 'Partido', 'Liga',
        'Apuesta Op1', 'Stake Op1', 'Resultado Op1', 'P/L Op1',
        'Apuesta Op4', 'Stake Op4', 'Resultado Op4', 'P/L Op4',
        'Dif P/L', 'Op1 Win', 'Op4 Win',
    ]
    fills_hdr = [fill('4472C4')] * 3 + [FILL_HDR1] * 4 + [FILL_HDR4] * 4 + [fill('7030A0')] * 3
    for ci, (h, f) in enumerate(zip(col_hdrs, fills_hdr), 1):
        c = ws.cell(4, ci, h)
        c.font = FONT_HDR; c.fill = f
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[4].height = 18
    ws.freeze_panes = 'A5'

    # --- Datos ---
    data_row = 5
    stats = {
        'op1': {'n': 0, 'g': 0, 'pl': 0.0, 'vol': 0.0},
        'op4': {'n': 0, 'g': 0, 'pl': 0.0, 'vol': 0.0},
    }

    for rd in datos:
        (id_p, fecha, local, visita, pais,
         p1, px, p2, po, pu,
         ap1x2, apou, stk1x2, stkou,
         c1, cx, c2, co, cu,
         estado, gl, gv, incert, auditoria,
         ap_shadow, stk_shadow, *_extra) = rd

        tiene_op1 = bool(stk1x2 and stk1x2 > 0 and ap1x2 and "[APOSTAR]" in str(ap1x2))
        tiene_op4 = bool(stk_shadow and stk_shadow > 0 and ap_shadow and "[APOSTAR]" in str(ap_shadow))
        if not tiene_op1 and not tiene_op4:
            continue

        fecha_disp = (fecha or "")[:10]
        for _fmt in ("%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                fecha_disp = datetime.strptime((fecha or "").strip(), _fmt).strftime("%d/%m/%Y")
                break
            except ValueError:
                continue

        # --- Op1 resultado ---
        res_op1 = determinar_resultado_entero(ap1x2, gl, gv) if tiene_op1 else 0
        if tiene_op1 and res_op1 != 0:
            cuota_op1 = cuota_1x2(ap1x2, c1, cx, c2) or 0
            pl_op1 = round(stk1x2 * (cuota_op1 - 1) if res_op1 == 1 else -stk1x2, 2)
            stats['op1']['n'] += 1
            stats['op1']['vol'] += stk1x2
            if res_op1 == 1:
                stats['op1']['g'] += 1
            stats['op1']['pl'] += pl_op1
            res_op1_str = "GANADA" if res_op1 == 1 else "PERDIDA"
            fill_res1 = FILL_GANADA_CELL if res_op1 == 1 else FILL_PERDIDA_CELL
            fill_pl1 = FILL_GANADA_CELL if pl_op1 > 0 else FILL_PERDIDA_CELL
        else:
            pl_op1 = None
            res_op1_str = "PENDIENTE" if gl is None else "-"
            fill_res1 = FILL_PEND_CELL if gl is None else FILL_BLANCO
            fill_pl1 = FILL_BLANCO

        # --- Op4 resultado ---
        res_op4 = determinar_resultado_entero(ap_shadow, gl, gv) if tiene_op4 else 0
        if tiene_op4 and res_op4 != 0:
            cuota_op4 = cuota_1x2(ap_shadow, c1, cx, c2) or 0
            pl_op4 = round(stk_shadow * (cuota_op4 - 1) if res_op4 == 1 else -stk_shadow, 2)
            stats['op4']['n'] += 1
            stats['op4']['vol'] += stk_shadow
            if res_op4 == 1:
                stats['op4']['g'] += 1
            stats['op4']['pl'] += pl_op4
            res_op4_str = "GANADA" if res_op4 == 1 else "PERDIDA"
            fill_res4 = FILL_GANADA_CELL if res_op4 == 1 else FILL_PERDIDA_CELL
            fill_pl4 = FILL_GANADA_CELL if pl_op4 > 0 else FILL_PERDIDA_CELL
        else:
            pl_op4 = None
            res_op4_str = "PENDIENTE" if gl is None else "-"
            fill_res4 = FILL_PEND_CELL if gl is None else FILL_BLANCO
            fill_pl4 = FILL_BLANCO

        pl_diff = round(pl_op1 - pl_op4, 2) if (pl_op1 is not None and pl_op4 is not None) else None
        fill_diff = (FILL_GANADA_CELL if pl_diff and pl_diff > 0 else
                     (FILL_PERDIDA_CELL if pl_diff and pl_diff < 0 else FILL_IGUAL)) \
                    if pl_diff is not None else FILL_BLANCO

        bg = FILL_BLANCO if data_row % 2 == 0 else FILL_NEUTRO_ROW

        def _dc(ci, val, fill_=None, fmt=None, left_align=False):
            cell = ws.cell(data_row, ci, val)
            cell.font = FONT_D
            cell.fill = fill_ or bg
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal='left' if left_align else 'center', vertical='center')
            if fmt:
                cell.number_format = fmt

        _dc(1, fecha_disp)
        _dc(2, f"{local} vs {visita}", left_align=True)
        _dc(3, pais)
        _dc(4, str(ap1x2 or "-"), left_align=True)
        _dc(5, stk1x2 if tiene_op1 else "-", fmt='#,##0.00')
        _dc(6, res_op1_str, fill_=fill_res1)
        _dc(7, pl_op1, fill_=fill_pl1, fmt='#,##0.00')
        _dc(8, str(ap_shadow or "-"), left_align=True)
        _dc(9, stk_shadow if tiene_op4 else "-", fmt='#,##0.00')
        _dc(10, res_op4_str, fill_=fill_res4)
        _dc(11, pl_op4, fill_=fill_pl4, fmt='#,##0.00')
        _dc(12, pl_diff, fill_=fill_diff, fmt='#,##0.00')
        _dc(13, "SI" if res_op1 == 1 else ("NO" if res_op1 == -1 else "-"),
            fill_=FILL_GANADA_CELL if res_op1 == 1 else
                 (FILL_PERDIDA_CELL if res_op1 == -1 else bg))
        _dc(14, "SI" if res_op4 == 1 else ("NO" if res_op4 == -1 else "-"),
            fill_=FILL_GANADA_CELL if res_op4 == 1 else
                 (FILL_PERDIDA_CELL if res_op4 == -1 else bg))

        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # --- Resumen KPI al pie ---
    data_row += 1
    o1, o4 = stats['op1'], stats['op4']
    n1, n4 = o1['n'], o4['n']

    ws.merge_cells(f'A{data_row}:N{data_row}')
    c = ws.cell(data_row, 1, "RESUMEN COMPARATIVO DE RENDIMIENTO")
    c.font = FONT_HDR; c.fill = fill('1F4E79')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[data_row].height = 18
    data_row += 1

    for ci, (h, f) in enumerate(zip(
        ['KPI', 'Opcion 1 (Activa)', '', 'Opcion 4 (Shadow)', '', 'Mejor'],
        [fill('4472C4'), FILL_HDR1, FILL_HDR1, FILL_HDR4, FILL_HDR4, fill('7030A0')],
    ), 1):
        if h:
            c = ws.cell(data_row, ci, h)
            c.font = FONT_HDR; c.fill = f
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = BORDER
    ws.row_dimensions[data_row].height = 16
    data_row += 1

    def _kpi(label, v1, v4, fmt='num', mayor_es_mejor=True, neutro=False):
        ws.row_dimensions[data_row].height = 17
        c = ws.cell(data_row, 1, label)
        c.font = FONT_KPI
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = BORDER

        def fmt_cell(ci, val, fill_):
            cell = ws.cell(data_row, ci, val)
            cell.font = FONT_D; cell.fill = fill_; cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if fmt == 'cur':  cell.number_format = '#,##0.00'
            elif fmt == 'pct': cell.number_format = '0.00%'
            elif fmt == 'int': cell.number_format = '0'

        if neutro or v1 == v4:
            f1 = f4 = FILL_IGUAL
            mejor_txt = "="
        elif mayor_es_mejor:
            f1 = FILL_MEJOR if v1 > v4 else FILL_PEOR
            f4 = FILL_MEJOR if v4 > v1 else FILL_PEOR
            mejor_txt = "Op1" if v1 > v4 else "Op4"
        else:
            f1 = FILL_MEJOR if v1 < v4 else FILL_PEOR
            f4 = FILL_MEJOR if v4 < v1 else FILL_PEOR
            mejor_txt = "Op1" if v1 < v4 else "Op4"

        fmt_cell(2, v1, f1)
        fmt_cell(3, v1, f1)
        fmt_cell(4, v4, f4)
        fmt_cell(5, v4, f4)
        c6 = ws.cell(data_row, 6, mejor_txt)
        c6.font = Font(name='Arial', bold=True, size=10)
        c6.fill = FILL_MEJOR if mejor_txt not in ("-", "=") else FILL_IGUAL
        c6.alignment = Alignment(horizontal='center', vertical='center')
        c6.border = BORDER

    yld1 = o1['pl'] / o1['vol'] if o1['vol'] else 0
    yld4 = o4['pl'] / o4['vol'] if o4['vol'] else 0
    hit1 = o1['g'] / n1 if n1 else 0
    hit4 = o4['g'] / n4 if n4 else 0

    _kpi("N apuestas liquidadas", n1, n4, 'int', neutro=True); data_row += 1
    _kpi("Ganadas", o1['g'], o4['g'], 'int');                   data_row += 1
    _kpi("Hit rate", hit1, hit4, 'pct');                        data_row += 1
    _kpi("Volumen apostado", o1['vol'], o4['vol'], 'cur', neutro=True); data_row += 1
    _kpi("P/L neto", o1['pl'], o4['pl'], 'cur');                data_row += 1
    _kpi("Yield", yld1, yld4, 'pct');                           data_row += 1

    data_row += 1
    ws.merge_cells(f'A{data_row}:N{data_row}')
    c = ws.cell(data_row, 1,
                "Verde = celda ganadora en ese KPI / partido   |   Rojo = celda perdedora   |   "
                "Amarillo en Resultado = pendiente de liquidar   |   "
                "Op1 activa desde V4.3: empates bloqueados + xG Margen O/U + Camino 2B (desacuerdo) + Camino 3 (alta conv.)")
    c.font = FONT_SUB
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
