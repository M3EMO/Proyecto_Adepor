"""
Hojas "Backtest" (principal con 1 fila por partido) y "Resumen" (por liga).

Extraido del motor_sincronizador.py monolitico en fase 4 (2026-04-21).
El main() del orquestador llama a:
  - poblar_backtest(wb, datos, bankroll) -> ws con todos los partidos + CF
  - stats = recolectar_stats_por_liga(datos)  (se alimenta en poblar_backtest)
  - crear_hoja_resumen(wb, stats, bankroll)
"""
from datetime import datetime, date as date_type

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from src.comun.resolucion import determinar_resultado_entero
from src.comun.calibracion_beta import obtener_coefs_beta, calibrar_probs
from src.persistencia.excel_estilos import (
    COL, CL, HEADERS, MAX_COL, COL_WIDTHS,
    FONT_HEADER, FONT_DATA, FILL_HEADER,
    FILL_GANADA, FILL_PERDIDA, FILL_PASAR, FILL_APOSTAR, FILL_PREDICCION,
    FILL_VERDE, FILL_AMARILLO, FILL_ROJO,
    ALIGN_CENTER, ALIGN_LEFT, BORDER_THIN, PAISES_CF,
)
from src.persistencia.excel_formulas import (
    f_apuesta_1x2, f_apuesta_ou, f_acierto, f_pl_neto, f_equity,
    f_brier, f_brier_casa, cuota_1x2, cuota_ou,
)


def poblar_backtest(wb, datos, bankroll):
    """Crea la hoja principal con un row por partido. Retorna dict stats_liga."""
    ws = wb.active
    ws.title = "Backtest"

    # Headers fila 1
    for col_idx, header_text in HEADERS.items():
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    for key, width in COL_WIDTHS.items():
        ws.column_dimensions[CL[key]].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(MAX_COL)}1'

    stats_liga = {}
    align_left_cols = {
        COL['partido'], COL['local'], COL['visita'],
        COL['ap1x2'], COL['apou'], COL['acierto'], COL['id'],
    }

    # Coefs beta para BS calibrado (display-only)
    coefs_beta = obtener_coefs_beta()

    for idx, row_data in enumerate(datos):
        r = idx + 2
        (id_p, fecha, local, visita, pais,
         p1, px, p2, po, pu,
         ap1x2, apou, stk1x2, stkou,
         c1, cx, c2, co, cu,
         estado, gl, gv, incert, auditoria,
         _ap_shadow, _stk_shadow, *_extra) = row_data

        # Fecha como date real (Excel ordena bien)
        fecha_val = fecha or ""
        for _fmt in ("%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                fecha_val = datetime.strptime((fecha or "").strip(), _fmt).date()
                break
            except ValueError:
                continue
        cell_fecha = ws.cell(r, COL['fecha'], fecha_val)
        cell_fecha.font = FONT_DATA
        if isinstance(fecha_val, date_type):
            cell_fecha.number_format = 'DD/MM/YYYY'

        ws.cell(r, COL['id'],      id_p).font             = FONT_DATA
        ws.cell(r, COL['partido'], f"{local} vs {visita}").font = FONT_DATA
        ws.cell(r, COL['local'],   local).font             = FONT_DATA
        ws.cell(r, COL['visita'],  visita).font            = FONT_DATA
        ws.cell(r, COL['liga'],    pais or "").font        = FONT_DATA

        for key, val in [('c1', c1), ('cx', cx), ('c2', c2), ('co', co), ('cu', cu)]:
            cell = ws.cell(r, COL[key], val if val and val > 0 else None)
            cell.font = FONT_DATA
            cell.number_format = '0.00'

        for key, val in [('p1', p1), ('px', px), ('p2', p2), ('po', po), ('pu', pu)]:
            cell = ws.cell(r, COL[key], val if val else None)
            cell.font = FONT_DATA
            cell.number_format = '0.0%'

        if gl is not None: ws.cell(r, COL['gl'], gl).font = FONT_DATA
        if gv is not None: ws.cell(r, COL['gv'], gv).font = FONT_DATA

        for key, val in [('stk1x2', stk1x2), ('stkou', stkou)]:
            cell = ws.cell(r, COL[key], val if val and val > 0 else 0)
            cell.font = FONT_DATA
            cell.number_format = '#,##0.00'

        if incert:
            cell = ws.cell(r, COL['incert'], incert)
            cell.font = FONT_DATA
            cell.number_format = '0.000'

            # Incert normalizada a [0, 1] con cap en 2.0, mostrada como %
            # 0% = muy predecible, 100% = maximamente impredecible
            incert_pct = min(float(incert) / 2.0, 1.0)
            cell = ws.cell(r, COL['incert_pct'], incert_pct)
            cell.font = FONT_DATA
            cell.number_format = '0%'

        ws.cell(r, COL['auditoria'], auditoria or "").font = FONT_DATA

        # Formulas
        ws.cell(r, COL['ap1x2'],   f_apuesta_1x2(r, ap1x2)).font = FONT_DATA
        ws.cell(r, COL['apou'],    f_apuesta_ou(r, apou)).font   = FONT_DATA
        ws.cell(r, COL['acierto'], f_acierto(r)).font            = FONT_DATA

        cell = ws.cell(r, COL['pl'], f_pl_neto(r))
        cell.font = FONT_DATA
        cell.number_format = '#,##0.00'

        cell = ws.cell(r, COL['equity'], f_equity(r, bankroll))
        cell.font = FONT_DATA
        cell.number_format = '#,##0.00'

        cell = ws.cell(r, COL['brier'], f_brier(r))
        cell.font = FONT_DATA
        cell.number_format = '0.0000'

        # BS calibrado: aplica beta-scaling a las probs y calcula valor numerico
        # (no formula) porque los coefs estan en Python, no en el workbook.
        if (isinstance(p1, (int, float)) and isinstance(px, (int, float)) and
                isinstance(p2, (int, float)) and p1 > 0 and
                gl is not None and gv is not None):
            q1, qx, q2 = calibrar_probs(p1, px, p2, coefs=coefs_beta)
            y1 = 1 if gl > gv else 0
            yx = 1 if gl == gv else 0
            y2 = 1 if gl < gv else 0
            bs_cal = (q1 - y1) ** 2 + (qx - yx) ** 2 + (q2 - y2) ** 2
            cell = ws.cell(r, COL['brier_cal'], round(bs_cal, 4))
            cell.font = FONT_DATA
            cell.number_format = '0.0000'

        cell = ws.cell(r, COL['brier_casa'], f_brier_casa(r))
        cell.font = FONT_DATA
        cell.number_format = '0.0000'

        # Bordes + alineacion
        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).border = BORDER_THIN
            ws.cell(r, c).alignment = ALIGN_LEFT if c in align_left_cols else ALIGN_CENTER

        # Stats por liga
        if pais:
            if pais not in stats_liga:
                stats_liga[pais] = {'apuestas': 0, 'ganadas': 0, 'perdidas': 0,
                                    'vol': 0.0, 'pl': 0.0}
            s = stats_liga[pais]
            if stk1x2 and stk1x2 > 0 and ap1x2:
                res = determinar_resultado_entero(ap1x2, gl, gv)
                if res != 0:
                    s['apuestas'] += 1
                    s['vol'] += stk1x2
                    if res == 1:
                        s['ganadas'] += 1
                        c_val = cuota_1x2(ap1x2, c1, cx, c2)
                        if c_val and c_val > 0:
                            s['pl'] += stk1x2 * (c_val - 1)
                    else:
                        s['perdidas'] += 1
                        s['pl'] -= stk1x2
            if stkou and stkou > 0 and apou:
                res = determinar_resultado_entero(apou, gl, gv)
                if res != 0:
                    s['apuestas'] += 1
                    s['vol'] += stkou
                    if res == 1:
                        s['ganadas'] += 1
                        c_val = cuota_ou(apou, co, cu)
                        if c_val and c_val > 0:
                            s['pl'] += stkou * (c_val - 1)
                    else:
                        s['perdidas'] += 1
                        s['pl'] -= stkou

    # Conditional Formatting
    _aplicar_conditional_formatting(ws, len(datos) + 1)

    return stats_liga


def _aplicar_conditional_formatting(ws, max_row):
    """Reglas de CF de la hoja Backtest: color por pais + colores de estado."""
    # 1. Filas por pais (A:R)
    rango_fila = f'A2:R{max_row}'
    for pais_cf, fill_cf in PAISES_CF:
        ws.conditional_formatting.add(rango_fila, FormulaRule(
            formula=[f'$K2="{pais_cf}"'], fill=fill_cf, stopIfTrue=False))

    # 2. Apuesta 1X2
    rango_s = f'{CL["ap1x2"]}2:{CL["ap1x2"]}{max_row}'
    for tag, fill_ in [("[GANADA]", FILL_GANADA), ("[PERDIDA]", FILL_PERDIDA),
                       ("[PASAR]",  FILL_PASAR),  ("[APOSTAR]", FILL_APOSTAR)]:
        ws.conditional_formatting.add(rango_s, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("{tag}",{CL["ap1x2"]}2))'],
            fill=fill_, stopIfTrue=True))

    # 3. Apuesta O/U
    rango_t = f'{CL["apou"]}2:{CL["apou"]}{max_row}'
    for tag, fill_ in [("[GANADA]", FILL_GANADA), ("[PERDIDA]", FILL_PERDIDA),
                       ("[PASAR]",  FILL_PASAR),  ("[APOSTAR]", FILL_APOSTAR)]:
        ws.conditional_formatting.add(rango_t, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("{tag}",{CL["apou"]}2))'],
            fill=fill_, stopIfTrue=True))

    # 4. P/L Neto
    rango_pl = f'{CL["pl"]}2:{CL["pl"]}{max_row}'
    ws.conditional_formatting.add(rango_pl, FormulaRule(
        formula=[f'{CL["pl"]}2>0'], fill=FILL_GANADA, stopIfTrue=True))
    ws.conditional_formatting.add(rango_pl, FormulaRule(
        formula=[f'{CL["pl"]}2<0'], fill=FILL_PERDIDA, stopIfTrue=True))

    # 5. Acierto
    rango_w = f'{CL["acierto"]}2:{CL["acierto"]}{max_row}'
    for tag, fill_ in [("[ACIERTO]", FILL_GANADA), ("[FALLO]", FILL_PERDIDA),
                       ("[PREDICCION]", FILL_PREDICCION), ("[PASAR]", FILL_PASAR)]:
        ws.conditional_formatting.add(rango_w, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("{tag}",{CL["acierto"]}2))'],
            fill=fill_, stopIfTrue=True))


def crear_hoja_resumen(wb, stats_liga, bankroll):
    """Hoja 'Resumen' con 1 fila por liga + TOTAL. CF por yield/P&L/acierto."""
    ws = wb.create_sheet("Resumen")
    res_headers = ['Liga', 'Apuestas', 'Ganadas', 'Perdidas',
                   '% Acierto', 'P/L Neto', 'Yield', 'Volumen']
    for i, h in enumerate(res_headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    row_r = 2
    total_ap = total_g = total_p = 0
    total_vol = total_pl = 0.0
    for liga, s in sorted(stats_liga.items()):
        pct = (s['ganadas'] / s['apuestas'] * 100) if s['apuestas'] > 0 else 0
        yld = (s['pl'] / s['vol'] * 100) if s['vol'] > 0 else 0
        ws.cell(row_r, 1, liga).font          = FONT_DATA
        ws.cell(row_r, 2, s['apuestas']).font = FONT_DATA
        ws.cell(row_r, 3, s['ganadas']).font  = FONT_DATA
        ws.cell(row_r, 4, s['perdidas']).font = FONT_DATA
        c = ws.cell(row_r, 5, pct / 100);   c.font = FONT_DATA; c.number_format = '0.0%'
        c = ws.cell(row_r, 6, round(s['pl'], 2)); c.font = FONT_DATA; c.number_format = '#,##0.00'
        c = ws.cell(row_r, 7, yld / 100);   c.font = FONT_DATA; c.number_format = '0.0%'
        c = ws.cell(row_r, 8, round(s['vol'], 2)); c.font = FONT_DATA; c.number_format = '#,##0.00'
        total_ap  += s['apuestas']
        total_g   += s['ganadas']
        total_p   += s['perdidas']
        total_vol += s['vol']
        total_pl  += s['pl']
        row_r += 1

    FONT_BOLD = Font(name='Arial', bold=True, size=10)
    ws.cell(row_r, 1, "TOTAL").font = FONT_BOLD
    ws.cell(row_r, 2, total_ap).font = FONT_BOLD
    ws.cell(row_r, 3, total_g).font = FONT_BOLD
    ws.cell(row_r, 4, total_p).font = FONT_BOLD
    c = ws.cell(row_r, 5, (total_g / total_ap) if total_ap > 0 else 0)
    c.font = FONT_BOLD; c.number_format = '0.0%'
    c = ws.cell(row_r, 6, round(total_pl, 2))
    c.font = FONT_BOLD; c.number_format = '#,##0.00'
    c = ws.cell(row_r, 7, (total_pl / total_vol) if total_vol > 0 else 0)
    c.font = FONT_BOLD; c.number_format = '0.0%'
    c = ws.cell(row_r, 8, round(total_vol, 2))
    c.font = FONT_BOLD; c.number_format = '#,##0.00'

    for i, w in enumerate([14, 10, 10, 10, 10, 14, 10, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # CF
    rng_yield = f'G2:G{row_r}'
    rng_pl    = f'F2:F{row_r}'
    rng_acie  = f'E2:E{row_r}'
    ws.conditional_formatting.add(rng_yield, FormulaRule(formula=['G2>0.05'], fill=FILL_VERDE,    stopIfTrue=True))
    ws.conditional_formatting.add(rng_yield, FormulaRule(formula=['G2>=0'],   fill=FILL_AMARILLO, stopIfTrue=True))
    ws.conditional_formatting.add(rng_yield, FormulaRule(formula=['G2<0'],    fill=FILL_ROJO,     stopIfTrue=True))
    ws.conditional_formatting.add(rng_pl,    FormulaRule(formula=['F2>0'],    fill=FILL_VERDE,    stopIfTrue=True))
    ws.conditional_formatting.add(rng_pl,    FormulaRule(formula=['F2<0'],    fill=FILL_ROJO,     stopIfTrue=True))
    ws.conditional_formatting.add(rng_acie,  FormulaRule(formula=['E2>=0.6'], fill=FILL_VERDE,    stopIfTrue=True))
    ws.conditional_formatting.add(rng_acie,  FormulaRule(formula=['E2>=0.5'], fill=FILL_AMARILLO, stopIfTrue=True))
    ws.conditional_formatting.add(rng_acie,  FormulaRule(formula=['E2<0.5'],  fill=FILL_ROJO,     stopIfTrue=True))

    ws.freeze_panes = 'A2'
    ws.cell(row_r + 2, 1,
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        Font(name='Arial', italic=True, size=9, color='888888')
    ws.cell(row_r + 3, 1, f"Bankroll base: ${bankroll:,.2f}").font = \
        Font(name='Arial', italic=True, size=9, color='888888')
