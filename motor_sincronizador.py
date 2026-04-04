import sqlite3
import os
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ==========================================
# MOTOR SINCRONIZADOR V9.1 (EXCEL LOCAL)
# Migrado de Google Sheets a Excel (.xlsx).
# Responsabilidad: Generar el archivo Excel de backtest con formulas vivas
# a partir de la base de datos SQLite (fuente de verdad).
# ==========================================

DB_NAME = 'fondo_quant.db'
EXCEL_FILE = 'Backtest_Modelo.xlsx'

# --- Mapa de columnas (1-indexed para openpyxl) ---
COL = {
    'fecha': 1, 'id': 2, 'partido': 3, 'local': 4, 'visita': 5,
    'c1': 6, 'cx': 7, 'c2': 8, 'co': 9, 'cu': 10,
    'liga': 11, 'p1': 12, 'px': 13, 'p2': 14, 'po': 15, 'pu': 16,
    'gl': 17, 'gv': 18,
    'ap1x2': 19, 'apou': 20, 'stk1x2': 21, 'stkou': 22,
    'acierto': 23, 'pl': 24, 'equity': 25, 'brier': 26,
    'incert': 27, 'auditoria': 28
}
HEADERS = {
    1: 'Fecha', 2: 'ID Partido', 3: 'Partido', 4: 'Local', 5: 'Visita',
    6: 'Cuota 1', 7: 'Cuota X', 8: 'Cuota 2', 9: 'Cuota +2.5', 10: 'Cuota -2.5',
    11: 'Liga', 12: 'Prob 1', 13: 'Prob X', 14: 'Prob 2', 15: 'Prob +2.5', 16: 'Prob -2.5',
    17: 'Goles L', 18: 'Goles V',
    19: 'Apuesta 1X2', 20: 'Apuesta O/U 2.5', 21: 'Stake 1X2', 22: 'Stake O/U 2.5',
    23: 'Acierto', 24: 'P/L Neto', 25: 'Equity Curve', 26: 'Brier Score',
    27: 'Incertidumbre', 28: 'Auditoria'
}
MAX_COL = max(COL.values())

# Letras de columna (precalculadas para formulas)
CL = {k: get_column_letter(v) for k, v in COL.items()}


# ==========================================================================
# HELPER: crea PatternFill con fgColor+bgColor identicos
# (necesario para que Excel renderice correctamente en CF/dxf)
# ==========================================================================
def _fill(hex_color):
    return PatternFill(patternType='solid', fgColor=hex_color, bgColor=hex_color)


# --- Estilos generales ---
FONT_HEADER  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
FONT_DATA    = Font(name='Arial', size=10)
FILL_HEADER  = _fill('1F4E79')
FILL_GANADA  = _fill('C6EFCE')
FILL_PERDIDA = _fill('FFC7CE')
FILL_PASAR   = _fill('FFEB9C')
FILL_APOSTAR = _fill('BDD7EE')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')
BORDER_THIN  = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),  bottom=Side(style='thin', color='D9D9D9')
)

# --- Colores de semaforo (dashboard) ---
FILL_VERDE   = _fill('C6EFCE')
FILL_AMARILLO= _fill('FFEB9C')
FILL_ROJO    = _fill('FFC7CE')
FILL_NEUTRO  = _fill('F2F2F2')

# --- Colores por pais (fila A:R en Backtest) ---
PAISES_CF = [
    ("Argentina", _fill('DEEAF1')),
    ("Brasil",    _fill('E2EFDA')),
    ("Noruega",   _fill('FFF2CC')),
    ("Turquia",   _fill('EAE0F0')),
    ("Inglaterra",_fill('FCE4D6')),
]

# --- Anchos de columna ---
COL_WIDTHS = {
    'fecha': 12, 'id': 32, 'partido': 30, 'local': 20, 'visita': 20,
    'c1': 9, 'cx': 9, 'c2': 9, 'co': 9, 'cu': 9,
    'liga': 12, 'p1': 9, 'px': 9, 'p2': 9, 'po': 9, 'pu': 9,
    'gl': 8, 'gv': 8,
    'ap1x2': 28, 'apou': 28, 'stk1x2': 13, 'stkou': 13,
    'acierto': 22, 'pl': 13, 'equity': 15, 'brier': 12,
    'incert': 13, 'auditoria': 11
}


# ==========================================================================
# GENERADORES DE FORMULAS EXCEL
# ==========================================================================

def f_apuesta_1x2(r, ap_text):
    """Formula de auto-liquidacion para Apuesta 1X2."""
    ap = str(ap_text or "")
    if "[APOSTAR]" not in ap:
        return ap
    g = f'{CL["gl"]}{r}'
    v = f'{CL["gv"]}{r}'
    vacio = f'OR({g}="",{v}="")'
    if "LOCAL" in ap:
        return f'=IF({vacio},"[APOSTAR] LOCAL",IF({g}>{v},"[GANADA] LOCAL","[PERDIDA] LOCAL"))'
    if "EMPATE" in ap:
        return f'=IF({vacio},"[APOSTAR] EMPATE",IF({g}={v},"[GANADA] EMPATE","[PERDIDA] EMPATE"))'
    if "VISITA" in ap:
        return f'=IF({vacio},"[APOSTAR] VISITA",IF({g}<{v},"[GANADA] VISITA","[PERDIDA] VISITA"))'
    return ap

def f_apuesta_ou(r, ap_text):
    """Formula de auto-liquidacion para Apuesta O/U."""
    ap = str(ap_text or "")
    if "[APOSTAR]" not in ap:
        return ap
    g = f'{CL["gl"]}{r}'
    v = f'{CL["gv"]}{r}'
    vacio = f'OR({g}="",{v}="")'
    total = f'({g}+{v})'
    if "OVER" in ap:
        return f'=IF({vacio},"[APOSTAR] OVER 2.5",IF({total}>2.5,"[GANADA] OVER 2.5","[PERDIDA] OVER 2.5"))'
    if "UNDER" in ap:
        return f'=IF({vacio},"[APOSTAR] UNDER 2.5",IF({total}<2.5,"[GANADA] UNDER 2.5","[PERDIDA] UNDER 2.5"))'
    return ap

def f_acierto(r):
    """Formula: Compara prediccion del modelo (prob mas alta) vs resultado real."""
    p1, px, p2 = f'{CL["p1"]}{r}', f'{CL["px"]}{r}', f'{CL["p2"]}{r}'
    gl, gv = f'{CL["gl"]}{r}', f'{CL["gv"]}{r}'
    mx = f'MAX({p1},{px},{p2})'
    md = f'MEDIAN({p1},{px},{p2})'
    pred  = f'IF({p1}={mx},"LOCAL",IF({px}={mx},"EMPATE","VISITA"))'
    res_l = f'IF({p1}={mx},IF({gl}>{gv},"[ACIERTO]","[FALLO]"),IF({px}={mx},IF({gl}={gv},"[ACIERTO]","[FALLO]"),IF({gl}<{gv},"[ACIERTO]","[FALLO]")))'
    return f'=IF({p1}="","",IF(({mx}-{md})>0.05,IF(OR({gl}="",{gv}=""),"[PREDICCION] "&{pred},{res_l}),"[PASAR] Margen Insuf"))'

def f_pl_neto(r):
    """Formula: P/L combinado de 1X2 + O/U."""
    s1, a1 = f'{CL["stk1x2"]}{r}', f'{CL["ap1x2"]}{r}'
    c1, cx, c2 = f'{CL["c1"]}{r}', f'{CL["cx"]}{r}', f'{CL["c2"]}{r}'
    so, ao  = f'{CL["stkou"]}{r}', f'{CL["apou"]}{r}'
    co, cu  = f'{CL["co"]}{r}', f'{CL["cu"]}{r}'
    pl1 = (f'IFERROR(IF({s1}=0,0,IF(ISNUMBER(SEARCH("[GANADA]",{a1})),'
           f'{s1}*(IF(ISNUMBER(SEARCH("LOCAL",{a1})),{c1},IF(ISNUMBER(SEARCH("EMPATE",{a1})),{cx},{c2}))-1),'
           f'IF(ISNUMBER(SEARCH("[PERDIDA]",{a1})),-{s1},0))),0)')
    plo = (f'IFERROR(IF({so}=0,0,IF(ISNUMBER(SEARCH("[GANADA]",{ao})),'
           f'{so}*(IF(ISNUMBER(SEARCH("OVER",{ao})),{co},{cu})-1),'
           f'IF(ISNUMBER(SEARCH("[PERDIDA]",{ao})),-{so},0))),0)')
    return f'={pl1}+{plo}'

def f_equity(r, bankroll):
    """Formula: Curva de equity acumulada."""
    pl = f'{CL["pl"]}{r}'
    eq_prev = f'{CL["equity"]}{r-1}'
    if r == 2:
        return f'={bankroll}+{pl}'
    return f'={eq_prev}+{pl}'

def f_brier(r):
    """Formula: Brier Score para mercado 1X2. Menor = mejor calibracion."""
    p1, px, p2 = f'{CL["p1"]}{r}', f'{CL["px"]}{r}', f'{CL["p2"]}{r}'
    gl, gv = f'{CL["gl"]}{r}', f'{CL["gv"]}{r}'
    return (f'=IF(OR({gl}="",{gv}=""),"",'
            f'({p1}-IF({gl}>{gv},1,0))^2+({px}-IF({gl}={gv},1,0))^2+({p2}-IF({gl}<{gv},1,0))^2)')


# ==========================================================================
# CALCULO DE METRICAS DASHBOARD
# ==========================================================================

def _resultado_1x2(ap_text, gl, gv):
    ap = str(ap_text or "")
    if "[GANADA]" in ap:  return  1
    if "[PERDIDA]" in ap: return -1
    if "[APOSTAR]" in ap and gl is not None and gv is not None:
        if "LOCAL"  in ap: return  1 if gl > gv else -1
        if "EMPATE" in ap: return  1 if gl == gv else -1
        if "VISITA" in ap: return  1 if gl < gv else -1
    return 0

def _resultado_ou(ap_text, gl, gv):
    ap = str(ap_text or "")
    if "[GANADA]" in ap:  return  1
    if "[PERDIDA]" in ap: return -1
    if "[APOSTAR]" in ap and gl is not None and gv is not None:
        total = gl + gv
        if "OVER"  in ap: return  1 if total > 2.5 else -1
        if "UNDER" in ap: return  1 if total < 2.5 else -1
    return 0

def _cuota_1x2(ap_text, c1, cx, c2):
    ap = str(ap_text or "")
    if "LOCAL"  in ap: return c1
    if "EMPATE" in ap: return cx
    if "VISITA" in ap: return c2
    return None

def _cuota_ou(ap_text, co, cu):
    ap = str(ap_text or "")
    if "OVER"  in ap: return co
    if "UNDER" in ap: return cu
    return None

def calcular_metricas_dashboard(datos, fraccion_kelly):
    bets_1x2, bets_ou = [], []
    bs_sis_list, bs_casa_list = [], []

    for row in datos:
        (id_p, fecha, local, visita, pais,
         p1, px, p2, po, pu,
         ap1x2, apou, stk1x2, stkou,
         c1, cx, c2, co, cu,
         estado, gl, gv, incert, auditoria) = row

        if gl is None or gv is None:
            continue

        if stk1x2 and stk1x2 > 0 and ap1x2:
            res = _resultado_1x2(ap1x2, gl, gv)
            if res != 0:
                cuota = _cuota_1x2(ap1x2, c1, cx, c2) or 0
                pl_val = stk1x2 * (cuota - 1) if res == 1 else -stk1x2
                bets_1x2.append({'res': res, 'stk': stk1x2, 'pl': pl_val})

        if stkou and stkou > 0 and apou:
            res = _resultado_ou(apou, gl, gv)
            if res != 0:
                cuota = _cuota_ou(apou, co, cu) or 0
                pl_val = stkou * (cuota - 1) if res == 1 else -stkou
                bets_ou.append({'res': res, 'stk': stkou, 'pl': pl_val})

        if p1 and px and p2:
            o1 = 1 if gl > gv else 0
            ox = 1 if gl == gv else 0
            o2 = 1 if gl < gv else 0
            bs_sis_list.append((p1-o1)**2 + (px-ox)**2 + (p2-o2)**2)
            if c1 and c1 > 0 and cx and cx > 0 and c2 and c2 > 0:
                r1, rx, r2 = 1/c1, 1/cx, 1/c2
                tot = r1 + rx + r2
                p1m, pxm, p2m = r1/tot, rx/tot, r2/tot
                bs_casa_list.append((p1m-o1)**2 + (pxm-ox)**2 + (p2m-o2)**2)

    def _grupo(bets):
        n = len(bets)
        if n == 0:
            return {'n':0,'pl':0.0,'vol':0.0,'yield':0.0,
                    'acierto_p':0.0,'acierto_dol':0.0,'t':0.0,'p':1.0}
        pl  = sum(b['pl']  for b in bets)
        vol = sum(b['stk'] for b in bets)
        ganadas = sum(1 for b in bets if b['res'] == 1)
        yld = pl / vol if vol > 0 else 0.0
        acierto_dol = (vol + pl) / vol if vol > 0 else 0.0
        if n >= 2:
            ys   = [b['pl']/b['stk'] for b in bets]
            mean_y = sum(ys)/n
            var  = sum((y-mean_y)**2 for y in ys)/(n-1)
            std  = math.sqrt(var) if var > 0 else 0.0
            t    = (mean_y / (std/math.sqrt(n))) if std > 0 else 0.0
            p_v  = math.erfc(abs(t)/math.sqrt(2))
        else:
            t, p_v = 0.0, 1.0
        return {'n':n,'pl':pl,'vol':vol,'yield':yld,
                'acierto_p': ganadas/n,
                'acierto_dol': acierto_dol,
                't': round(t, 4), 'p': round(p_v, 4)}

    m_all = _grupo(bets_1x2 + bets_ou)
    m_1x2 = _grupo(bets_1x2)
    m_ou  = _grupo(bets_ou)

    bs_sis  = sum(bs_sis_list)/len(bs_sis_list)   if bs_sis_list  else 0.0
    bs_casa = sum(bs_casa_list)/len(bs_casa_list)  if bs_casa_list else 0.0

    return {
        'total': m_all, '1x2': m_1x2, 'ou': m_ou,
        'bs_sis':  bs_sis,
        'bs_casa': bs_casa,
        'bs_glob': bs_sis - bs_casa,
        'fraccion_kelly': fraccion_kelly,
    }


# ==========================================================================
# SEMAFORO: devuelve fill segun umbrales
# ==========================================================================
def _semaforo(valor, bueno, malo, mayor_es_mejor=True):
    """
    mayor_es_mejor=True  → valor >= bueno: verde | malo..bueno: amarillo | < malo: rojo
    mayor_es_mejor=False → valor <= bueno: verde | bueno..malo: amarillo | > malo: rojo
    Devuelve None si valor es '—' o None.
    """
    if valor is None or valor == '—':
        return FILL_NEUTRO
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return FILL_NEUTRO
    if mayor_es_mejor:
        if v >= bueno: return FILL_VERDE
        if v >= malo:  return FILL_AMARILLO
        return FILL_ROJO
    else:
        if v <= bueno: return FILL_VERDE
        if v <= malo:  return FILL_AMARILLO
        return FILL_ROJO


# ==========================================================================
# HOJA DASHBOARD
# ==========================================================================

def crear_hoja_dashboard(wb, metricas, bankroll):
    ws = wb.create_sheet("Dashboard", 0)

    FONT_TITLE  = Font(name='Arial', bold=True, color='FFFFFF', size=13)
    FONT_SEC    = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    FONT_KPI    = Font(name='Arial', bold=True, size=10)
    FONT_VAL    = Font(name='Arial', size=10)
    FONT_SUB    = Font(name='Arial', italic=True, size=9, color='595959')
    FILL_TITLE  = _fill('1F4E79')
    FILL_SEC    = _fill('4472C4')
    FILL_HDR_COL= _fill('2E75B6')
    BORDER_DB   = Border(
        left=Side(style='thin', color='9DC3E6'),
        right=Side(style='thin', color='9DC3E6'),
        top=Side(style='thin', color='9DC3E6'),
        bottom=Side(style='thin', color='9DC3E6')
    )

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # ---- Titulo ----
    ws.merge_cells('A1:D1')
    c = ws.cell(1, 1, "DASHBOARD DE RENDIMIENTO")
    c.font = FONT_TITLE; c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    c = ws.cell(2, 1, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Bankroll: ${bankroll:,.2f}")
    c.font = FONT_SUB; c.fill = _fill('D6E4F0')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 15

    # ---- Cabecera de columnas ----
    for ci, h in enumerate(['Metrica', 'Total', '1X2', 'O/U'], 1):
        c = ws.cell(3, ci, h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = FILL_HDR_COL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER_DB
    ws.row_dimensions[3].height = 20

    m = metricas
    t_all, t_1x2, t_ou = m['total'], m['1x2'], m['ou']
    row = 4

    # ---- Helpers internos ----
    def _sep(titulo):
        nonlocal row
        ws.merge_cells(f'A{row}:D{row}')
        c = ws.cell(row, 1, titulo)
        c.font = FONT_SEC; c.fill = FILL_SEC
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = BORDER_DB
        ws.row_dimensions[row].height = 14
        row += 1

    def _fila(metrica, vals, fmts, fills):
        """
        vals  = (total, 1x2, ou)   — puede ser '—' para N/A
        fmts  = (fmt_total, fmt_1x2, fmt_ou)
        fills = (fill_total, fill_1x2, fill_ou)  — PatternFill o None
        """
        nonlocal row
        is_alt = (row % 2 == 0)
        bg_base = _fill('EBF3FB') if is_alt else _fill('FFFFFF')
        ws.row_dimensions[row].height = 18

        c = ws.cell(row, 1, metrica)
        c.font = FONT_KPI; c.fill = bg_base; c.border = BORDER_DB
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        for ci, (val, fmt, fill) in enumerate(zip(vals, fmts, fills), 2):
            cell = ws.cell(row, ci, val)
            cell.font = FONT_VAL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_DB
            cell.fill = fill if fill else bg_base
            if fmt == 'pct':   cell.number_format = '0.00%'
            elif fmt == 'cur': cell.number_format = '#,##0.00'
            elif fmt == 'd4':  cell.number_format = '0.0000'
            elif fmt == 'd2':  cell.number_format = '0.00'
            elif fmt == 'int': cell.number_format = '0'
        row += 1

    NA = '—'
    n_all = t_all['n']

    # ==== BLOQUE 1: Financiero ====
    _sep("  RESULTADOS FINANCIEROS")

    _fila("Ganancia neta",
          (t_all['pl'],   t_1x2['pl'],   t_ou['pl']),
          ('cur','cur','cur'),
          (_semaforo(t_all['pl'],  0, 0),
           _semaforo(t_1x2['pl'], 0, 0),
           _semaforo(t_ou['pl'],  0, 0)))

    # Yield: >5% bueno, 0-5% aceptable, <0% malo
    _fila("Yield",
          (t_all['yield'],  t_1x2['yield'],  t_ou['yield']),
          ('pct','pct','pct'),
          (_semaforo(t_all['yield'],  0.05, 0),
           _semaforo(t_1x2['yield'], 0.05, 0),
           _semaforo(t_ou['yield'],  0.05, 0)))

    _fila("Volumen apostado",
          (t_all['vol'],   t_1x2['vol'],   t_ou['vol']),
          ('cur','cur','cur'),
          (FILL_NEUTRO, FILL_NEUTRO, FILL_NEUTRO))

    _fila("N apuestas liquidadas",
          (t_all['n'],  t_1x2['n'],  t_ou['n']),
          ('int','int','int'),
          (FILL_NEUTRO, FILL_NEUTRO, FILL_NEUTRO))

    # ==== BLOQUE 2: Acierto ====
    _sep("  TASA DE ACIERTO")

    _fila("% Acierto P  (hit rate por conteo)",
          (t_all['acierto_p'],  t_1x2['acierto_p'],  t_ou['acierto_p']),
          ('pct','pct','pct'),
          (_semaforo(t_all['acierto_p'],  0.55, 0.45),
           _semaforo(t_1x2['acierto_p'], 0.55, 0.45),
           _semaforo(t_ou['acierto_p'],  0.55, 0.45)))

    _fila("% Acierto $  (ROI, capital recuperado)",
          (t_all['acierto_dol'],  t_1x2['acierto_dol'],  t_ou['acierto_dol']),
          ('pct','pct','pct'),
          (_semaforo(t_all['acierto_dol'],  1.03, 0.97),
           _semaforo(t_1x2['acierto_dol'], 1.03, 0.97),
           _semaforo(t_ou['acierto_dol'],  1.03, 0.97)))

    won_all  = round(t_all['acierto_p']  * t_all['n'])
    won_1x2  = round(t_1x2['acierto_p'] * t_1x2['n'])
    won_ou   = round(t_ou['acierto_p']  * t_ou['n'])
    acp_all  = won_all  / t_all['n']  if t_all['n']  > 0 else 0
    acp_1x2  = won_1x2  / t_1x2['n'] if t_1x2['n'] > 0 else 0
    acp_ou   = won_ou   / t_ou['n']  if t_ou['n']  > 0 else 0
    _fila("% Acierto all  (ganadas / total)",
          (acp_all, acp_1x2, acp_ou),
          ('pct','pct','pct'),
          (_semaforo(acp_all, 0.55, 0.45),
           _semaforo(acp_1x2, 0.55, 0.45),
           _semaforo(acp_ou,  0.55, 0.45)))

    # ==== BLOQUE 3: Estadistica ====
    _sep("  ESTADISTICA INFERENCIAL")

    _fila("T-score",
          (t_all['t'],  t_1x2['t'],  t_ou['t']),
          ('d2','d2','d2'),
          (_semaforo(abs(t_all['t']),  2.0, 1.0),
           _semaforo(abs(t_1x2['t']), 2.0, 1.0),
           _semaforo(abs(t_ou['t']),  2.0, 1.0)))

    _fila("P-Value  (two-tailed, <0.05 = significativo)",
          (t_all['p'],  t_1x2['p'],  t_ou['p']),
          ('d4','d4','d4'),
          (_semaforo(t_all['p'],  0.05, 0.10, mayor_es_mejor=False),
           _semaforo(t_1x2['p'], 0.05, 0.10, mayor_es_mejor=False),
           _semaforo(t_ou['p'],  0.05, 0.10, mayor_es_mejor=False)))

    fk = m['fraccion_kelly']
    _fila("Fraccion Kelly",
          (fk, fk, NA),
          ('pct','pct',''),
          (FILL_NEUTRO, FILL_NEUTRO, FILL_NEUTRO))

    # ==== BLOQUE 4: Calibracion ====
    _sep("  CALIBRACION DEL MODELO  (Brier Score — menor es mejor)")

    bs_s  = m['bs_sis']
    bs_c  = m['bs_casa']
    bs_g  = m['bs_glob']

    # BS rango 0-2 (3 resultados posibles). Aleatorio puro ≈ 0.667
    # Verde: < 0.50 (mejor que azar), Amarillo: 0.50-0.65, Rojo: >= 0.65
    _fila("Brier Score sistema  (Dixon-Coles)",
          (bs_s, bs_s, NA), ('d4','d4',''),
          (_semaforo(bs_s, 0.50, 0.65, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    _fila("Promedio BS  (media por partido)",
          (bs_s, bs_s, NA), ('d4','d4',''),
          (_semaforo(bs_s, 0.50, 0.65, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    # BS casa: referencia neutra, pero coloreamos igual para comparar
    _fila("BS casa  (Pinnacle / bet365 referencia)",
          (bs_c, bs_c, NA), ('d4','d4',''),
          (_semaforo(bs_c, 0.50, 0.65, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    # BS global: diferencia sistema - casa. Negativo = nuestro modelo es mejor
    _fila("BS global  (sistema - casa, negativo = mejor)",
          (bs_g, bs_g, NA), ('d4','d4',''),
          (_semaforo(bs_g, -0.02, 0.02, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    # ---- Nota al pie ----
    ws.merge_cells(f'A{row+1}:D{row+1}')
    c = ws.cell(row+1, 1,
        "Umbrales — Yield: >5%=verde / 0-5%=amarillo / <0%=rojo  |  "
        "BS (rango 0-2, aleatorio≈0.667): <0.50=verde / 0.50-0.65=amarillo / >0.65=rojo  |  "
        "P-Value: <0.05=verde / 0.05-0.10=amarillo / >0.10=rojo  |  BS global neg. = modelo supera mercado")
    c.font = Font(name='Arial', italic=True, size=8, color='595959')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws.freeze_panes = 'A4'


# ==========================================================================
# FUNCION PRINCIPAL
# ==========================================================================

def main():
    print("[SISTEMA] Iniciando Motor Sincronizador V9.1 (Excel Local)...")

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # --- Resurrection: partidos Liquidados sin goles vuelven a Calculado ---
    cursor.execute("""
        SELECT id_partido FROM partidos_backtest
        WHERE estado = 'Liquidado' AND (goles_l IS NULL OR goles_v IS NULL)
    """)
    resurrecciones = [r[0] for r in cursor.fetchall()]
    if resurrecciones:
        cursor.executemany(
            "UPDATE partidos_backtest SET estado = 'Calculado' WHERE id_partido = ?",
            [(i,) for i in resurrecciones]
        )
        conn.commit()
        print(f"[INFO] {len(resurrecciones)} partidos resucitados (Liquidado sin goles -> Calculado).")

    # --- Bankroll ---
    try:
        cursor.execute("SELECT valor FROM configuracion WHERE clave = 'bankroll'")
        BANKROLL = float(cursor.fetchone()[0])
    except (TypeError, IndexError):
        BANKROLL = 100000.00

    # --- Fraccion Kelly ---
    try:
        cursor.execute("SELECT valor FROM configuracion WHERE clave = 'fraccion_kelly'")
        row_fk = cursor.fetchone()
        FRACCION_KELLY = float(row_fk[0]) if row_fk else 0.50
    except Exception:
        FRACCION_KELLY = 0.50

    # --- Datos principales ---
    cursor.execute("""
        SELECT id_partido, fecha, local, visita, pais,
               prob_1, prob_x, prob_2, prob_o25, prob_u25,
               apuesta_1x2, apuesta_ou, stake_1x2, stake_ou,
               cuota_1, cuota_x, cuota_2, cuota_o25, cuota_u25,
               estado, goles_l, goles_v, incertidumbre, auditoria
        FROM partidos_backtest
        WHERE estado IN ('Calculado', 'Liquidado')
        ORDER BY id_partido ASC
    """)
    datos = cursor.fetchall()
    conn.close()

    if not datos:
        print("[INFO] No hay partidos para sincronizar.")
        return

    print(f"[INFO] {len(datos)} partidos a sincronizar. Bankroll: ${BANKROLL:,.2f}")

    # --- Crear workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest"

    # --- Headers ---
    for col_idx, header_text in HEADERS.items():
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    # --- Anchos de columna ---
    for key, width in COL_WIDTHS.items():
        ws.column_dimensions[CL[key]].width = width

    # --- Freeze y autofiltro ---
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(MAX_COL)}1'

    # --- Estadisticas por liga (para hoja Resumen) ---
    stats_liga = {}

    # --- Escribir datos ---
    for idx, row_data in enumerate(datos):
        r = idx + 2
        (id_p, fecha, local, visita, pais,
         p1, px, p2, po, pu,
         ap1x2, apou, stk1x2, stkou,
         c1, cx, c2, co, cu,
         estado, gl, gv, incert, auditoria) = row_data

        ws.cell(r, COL['fecha'],   (fecha.split(" ")[0] if fecha else "")).font = FONT_DATA
        ws.cell(r, COL['id'],      id_p).font = FONT_DATA
        ws.cell(r, COL['partido'], f"{local} vs {visita}").font = FONT_DATA
        ws.cell(r, COL['local'],   local).font = FONT_DATA
        ws.cell(r, COL['visita'],  visita).font = FONT_DATA
        ws.cell(r, COL['liga'],    pais or "").font = FONT_DATA

        for key, val in [('c1',c1),('cx',cx),('c2',c2),('co',co),('cu',cu)]:
            cell = ws.cell(r, COL[key], val if val and val > 0 else None)
            cell.font = FONT_DATA; cell.number_format = '0.00'

        for key, val in [('p1',p1),('px',px),('p2',p2),('po',po),('pu',pu)]:
            cell = ws.cell(r, COL[key], val if val else None)
            cell.font = FONT_DATA; cell.number_format = '0.0%'

        if gl is not None: ws.cell(r, COL['gl'], gl).font = FONT_DATA
        if gv is not None: ws.cell(r, COL['gv'], gv).font = FONT_DATA

        for key, val in [('stk1x2', stk1x2), ('stkou', stkou)]:
            cell = ws.cell(r, COL[key], val if val and val > 0 else 0)
            cell.font = FONT_DATA; cell.number_format = '#,##0.00'

        if incert:
            cell = ws.cell(r, COL['incert'], incert)
            cell.font = FONT_DATA; cell.number_format = '0.000'

        ws.cell(r, COL['auditoria'], auditoria or "").font = FONT_DATA

        ws.cell(r, COL['ap1x2'],  f_apuesta_1x2(r, ap1x2)).font = FONT_DATA
        ws.cell(r, COL['apou'],   f_apuesta_ou(r,  apou)).font  = FONT_DATA
        ws.cell(r, COL['acierto'],f_acierto(r)).font             = FONT_DATA
        ws.cell(r, COL['pl'],     f_pl_neto(r)).font             = FONT_DATA
        ws.cell(r, COL['pl']).number_format = '#,##0.00'
        ws.cell(r, COL['equity'], f_equity(r, BANKROLL)).font    = FONT_DATA
        ws.cell(r, COL['equity']).number_format = '#,##0.00'
        ws.cell(r, COL['brier'],  f_brier(r)).font               = FONT_DATA
        ws.cell(r, COL['brier']).number_format = '0.0000'

        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).border = BORDER_THIN
            ws.cell(r, c).alignment = (
                ALIGN_LEFT if c in [COL['partido'], COL['local'], COL['visita'],
                                    COL['ap1x2'], COL['apou'], COL['acierto'], COL['id']]
                else ALIGN_CENTER
            )

        if pais:
            if pais not in stats_liga:
                stats_liga[pais] = {'apuestas':0,'ganadas':0,'perdidas':0,'vol':0.0,'pl':0.0}
            s = stats_liga[pais]
            for ap, stk, cuotas_dict in [
                (str(ap1x2 or ""), stk1x2 or 0, {'LOCAL':c1,'EMPATE':cx,'VISITA':c2}),
                (str(apou  or ""), stkou  or 0, {'OVER':co,'UNDER':cu})
            ]:
                if stk > 0 and ("[GANADA]" in ap or "[PERDIDA]" in ap):
                    s['apuestas'] += 1
                    s['vol'] += stk
                    if "[GANADA]" in ap:
                        s['ganadas'] += 1
                        for k, v in cuotas_dict.items():
                            if k in ap and v and v > 0:
                                s['pl'] += stk * (v - 1)
                    else:
                        s['perdidas'] += 1
                        s['pl'] -= stk

    max_row = len(datos) + 1

    # ==========================================================================
    # CONDITIONAL FORMATTING — regenerado en cada ejecucion (no se pierde)
    # ==========================================================================

    # 1. Coloreado de filas por pais (A:R)
    rango_fila = f'A2:R{max_row}'
    for pais_cf, fill_cf in PAISES_CF:
        ws.conditional_formatting.add(rango_fila, FormulaRule(
            formula=[f'$K2="{pais_cf}"'], fill=fill_cf, stopIfTrue=False))

    # 2. Columna S — Apuesta 1X2
    rango_s = f'{CL["ap1x2"]}2:{CL["ap1x2"]}{max_row}'
    ws.conditional_formatting.add(rango_s, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[GANADA]",{CL["ap1x2"]}2))'],  fill=FILL_GANADA,  stopIfTrue=True))
    ws.conditional_formatting.add(rango_s, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[PERDIDA]",{CL["ap1x2"]}2))'], fill=FILL_PERDIDA, stopIfTrue=True))
    ws.conditional_formatting.add(rango_s, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[PASAR]",{CL["ap1x2"]}2))'],   fill=FILL_PASAR,   stopIfTrue=True))
    ws.conditional_formatting.add(rango_s, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[APOSTAR]",{CL["ap1x2"]}2))'], fill=FILL_APOSTAR, stopIfTrue=True))

    # 3. Columna T — Apuesta O/U
    rango_t = f'{CL["apou"]}2:{CL["apou"]}{max_row}'
    ws.conditional_formatting.add(rango_t, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[GANADA]",{CL["apou"]}2))'],  fill=FILL_GANADA,  stopIfTrue=True))
    ws.conditional_formatting.add(rango_t, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[PERDIDA]",{CL["apou"]}2))'], fill=FILL_PERDIDA, stopIfTrue=True))
    ws.conditional_formatting.add(rango_t, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[PASAR]",{CL["apou"]}2))'],   fill=FILL_PASAR,   stopIfTrue=True))
    ws.conditional_formatting.add(rango_t, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[APOSTAR]",{CL["apou"]}2))'], fill=FILL_APOSTAR, stopIfTrue=True))

    # 4. Columna W — Acierto ([PASAR] = amarillo)
    rango_w = f'{CL["acierto"]}2:{CL["acierto"]}{max_row}'
    ws.conditional_formatting.add(rango_w, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("[PASAR]",{CL["acierto"]}2))'], fill=FILL_PASAR, stopIfTrue=True))

    # ==========================================================================
    # DASHBOARD + RESUMEN
    # ==========================================================================
    metricas = calcular_metricas_dashboard(datos, FRACCION_KELLY)
    crear_hoja_dashboard(wb, metricas, BANKROLL)

    # --- Hoja de Resumen ---
    ws2 = wb.create_sheet("Resumen")
    res_headers = ['Liga','Apuestas','Ganadas','Perdidas','% Acierto','P/L Neto','Yield','Volumen']
    for i, h in enumerate(res_headers, 1):
        cell = ws2.cell(1, i, h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = ALIGN_CENTER

    row_r = 2
    total_ap, total_g, total_p, total_vol, total_pl = 0, 0, 0, 0.0, 0.0
    for liga, s in sorted(stats_liga.items()):
        pct = (s['ganadas'] / s['apuestas'] * 100) if s['apuestas'] > 0 else 0
        yld = (s['pl'] / s['vol'] * 100)           if s['vol']      > 0 else 0
        ws2.cell(row_r, 1, liga).font           = FONT_DATA
        ws2.cell(row_r, 2, s['apuestas']).font  = FONT_DATA
        ws2.cell(row_r, 3, s['ganadas']).font   = FONT_DATA
        ws2.cell(row_r, 4, s['perdidas']).font  = FONT_DATA
        c = ws2.cell(row_r, 5, pct/100); c.font = FONT_DATA; c.number_format = '0.0%'
        c = ws2.cell(row_r, 6, round(s['pl'],2)); c.font = FONT_DATA; c.number_format = '#,##0.00'
        c = ws2.cell(row_r, 7, yld/100); c.font = FONT_DATA; c.number_format = '0.0%'
        c = ws2.cell(row_r, 8, round(s['vol'],2)); c.font = FONT_DATA; c.number_format = '#,##0.00'
        total_ap += s['apuestas']; total_g += s['ganadas']
        total_p  += s['perdidas']; total_vol += s['vol']; total_pl += s['pl']
        row_r += 1

    FONT_BOLD = Font(name='Arial', bold=True, size=10)
    ws2.cell(row_r, 1, "TOTAL").font = FONT_BOLD
    ws2.cell(row_r, 2, total_ap).font = FONT_BOLD
    ws2.cell(row_r, 3, total_g).font  = FONT_BOLD
    ws2.cell(row_r, 4, total_p).font  = FONT_BOLD
    c = ws2.cell(row_r, 5, (total_g/total_ap) if total_ap > 0 else 0)
    c.font = FONT_BOLD; c.number_format = '0.0%'
    c = ws2.cell(row_r, 6, round(total_pl, 2))
    c.font = FONT_BOLD; c.number_format = '#,##0.00'
    c = ws2.cell(row_r, 7, (total_pl/total_vol) if total_vol > 0 else 0)
    c.font = FONT_BOLD; c.number_format = '0.0%'
    c = ws2.cell(row_r, 8, round(total_vol, 2))
    c.font = FONT_BOLD; c.number_format = '#,##0.00'

    for i, w in enumerate([14,10,10,10,10,14,10,14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'
    ws2.cell(row_r+2, 1, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        Font(name='Arial', italic=True, size=9, color='888888')
    ws2.cell(row_r+3, 1, f"Bankroll base: ${BANKROLL:,.2f}").font = \
        Font(name='Arial', italic=True, size=9, color='888888')

    # --- Guardar ---
    wb.calculation.fullCalcOnLoad = True
    wb.save(EXCEL_FILE)
    print(f"[EXITO] Excel generado: {os.path.abspath(EXCEL_FILE)}")
    print(f"[INFO] {len(datos)} partidos escritos. {len(stats_liga)} ligas resumidas.")
    print("[SISTEMA] Motor Sincronizador V9.1 ha finalizado su ejecucion.")

if __name__ == "__main__":
    main()
