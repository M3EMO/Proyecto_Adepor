"""
Estilos y schema de columnas para el Excel del motor_sincronizador.

Extraido del motor_sincronizador.py monolitico (V9.2) en la modularizacion
fase 4 (2026-04-21). Aqui vive TODO lo cosmetico: fills, fonts, alignments,
colores por liga, schema de columnas. Los demas modulos (excel_formulas,
excel_metricas, excel_hoja_*) importan desde aqui.
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ==========================================================================
# SCHEMA DE COLUMNAS (hoja Backtest)
# ==========================================================================
COL = {
    'fecha': 1, 'id': 2, 'partido': 3, 'local': 4, 'visita': 5,
    'c1': 6, 'cx': 7, 'c2': 8, 'co': 9, 'cu': 10,
    'liga': 11, 'p1': 12, 'px': 13, 'p2': 14, 'po': 15, 'pu': 16,
    'gl': 17, 'gv': 18,
    'ap1x2': 19, 'apou': 20, 'stk1x2': 21, 'stkou': 22,
    'acierto': 23, 'pl': 24, 'equity': 25,
    'brier': 26, 'brier_cal': 27, 'brier_casa': 28,
    'incert': 29, 'incert_pct': 30, 'auditoria': 31,
}

HEADERS = {
    1: 'Fecha', 2: 'ID Partido', 3: 'Partido', 4: 'Local', 5: 'Visita',
    6: 'Cuota 1', 7: 'Cuota X', 8: 'Cuota 2', 9: 'Cuota +2.5', 10: 'Cuota -2.5',
    11: 'Liga', 12: 'Prob 1', 13: 'Prob X', 14: 'Prob 2', 15: 'Prob +2.5', 16: 'Prob -2.5',
    17: 'Goles L', 18: 'Goles V',
    19: 'Apuesta 1X2', 20: 'Apuesta O/U 2.5', 21: 'Stake 1X2', 22: 'Stake O/U 2.5',
    23: 'Acierto', 24: 'P/L Neto', 25: 'Equity Curve',
    26: 'BS Sistema', 27: 'BS Calibrado', 28: 'BS Casa',
    29: 'Incertidumbre', 30: 'Incert %', 31: 'Auditoria',
}

MAX_COL = max(COL.values())
CL = {k: get_column_letter(v) for k, v in COL.items()}

COL_WIDTHS = {
    'fecha': 12, 'id': 32, 'partido': 30, 'local': 20, 'visita': 20,
    'c1': 9, 'cx': 9, 'c2': 9, 'co': 9, 'cu': 9,
    'liga': 12, 'p1': 9, 'px': 9, 'p2': 9, 'po': 9, 'pu': 9,
    'gl': 8, 'gv': 8,
    'ap1x2': 28, 'apou': 28, 'stk1x2': 13, 'stkou': 13,
    'acierto': 22, 'pl': 13, 'equity': 15,
    'brier': 12, 'brier_cal': 12, 'brier_casa': 12,
    'incert': 13, 'incert_pct': 10, 'auditoria': 11,
}


# ==========================================================================
# HELPER FILL (fgColor + bgColor para que Excel Conditional Formatting renderice)
# ==========================================================================
def fill(hex_color):
    return PatternFill(patternType='solid', fgColor=hex_color, bgColor=hex_color)


# ==========================================================================
# ESTILOS GENERALES (hoja Backtest)
# ==========================================================================
FONT_HEADER     = Font(name='Arial', bold=True, color='FFFFFF', size=10)
FONT_DATA       = Font(name='Arial', size=10)
FILL_HEADER     = fill('1F4E79')
FILL_GANADA     = fill('C6EFCE')
FILL_PERDIDA    = fill('FFC7CE')
FILL_PASAR      = fill('FFEB9C')
FILL_APOSTAR    = fill('BDD7EE')
FILL_PREDICCION = fill('9DC3E6')
ALIGN_CENTER    = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT      = Alignment(horizontal='left',   vertical='center')
BORDER_THIN     = Border(
    left=Side(style='thin',   color='D9D9D9'), right=Side(style='thin',  color='D9D9D9'),
    top=Side(style='thin',    color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'),
)

# ==========================================================================
# COLORES SEMAFORO (Dashboard)
# ==========================================================================
FILL_VERDE    = fill('C6EFCE')
FILL_AMARILLO = fill('FFEB9C')
FILL_ROJO     = fill('FFC7CE')
FILL_NEUTRO   = fill('F2F2F2')

# ==========================================================================
# COLORES POR PAIS (Conditional Formatting fila A:R en Backtest)
# Paleta pastel diferenciable. Evitar repetir tonos entre paises vecinos.
# Argentina/Uruguay comparten familia "celeste" pero con intensidad distinta
# para diferenciarlos (Argentina mas azul, Uruguay mas verde agua).
# ==========================================================================
PAISES_CF = [
    # --- Sudamericanas ---
    ("Argentina",  fill('DEEAF1')),   # celeste claro
    ("Brasil",     fill('E2EFDA')),   # verde lima claro
    ("Bolivia",    fill('F3E5F5')),   # lavanda
    ("Chile",      fill('F5CBA7')),   # salmon claro
    ("Colombia",   fill('FCF3CF')),   # amarillo crema (bandera)
    ("Ecuador",    fill('FAD7A0')),   # durazno (bandera)
    ("Peru",       fill('F9E79F')),   # mostaza suave (bandera)
    ("Uruguay",    fill('A9DFBF')),   # verde agua (bandera celeste distinta a ARG)
    ("Venezuela",  fill('F5B7B1')),   # rosa coral (bandera)
    # --- Europeas ---
    ("Noruega",    fill('FFF2CC')),   # amarillo pastel
    ("Turquia",    fill('EAE0F0')),   # violeta claro
    ("Inglaterra", fill('FCE4D6')),   # naranja durazno
    ("Espana",     fill('F8CECC')),   # rosa coral (bandera roja)
    ("Italia",     fill('D5E8D4')),   # verde menta (bandera)
    ("Alemania",   fill('FFF4E6')),   # beige dorado (bandera amarilla/negra)
    ("Francia",    fill('DAE8FC')),   # azul periwinkle (bandera azul)
]
