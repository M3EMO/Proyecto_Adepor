"""
Hoja "Si Hubiera" — resimulacion de todo el historico de partidos liquidados
aplicando las reglas actuales (Fase 3.3.5).

Responde: "cuanto habria rendido el sistema si siempre hubiera apostado con los
criterios de hoy". ES BACKTESTING IN-SAMPLE (las reglas fueron calibradas sobre
exactamente estos datos) — sirve para visualizar, no para decidir LIVE.

Estructura:
  1. Titulo + advertencia
  2. KPIs globales (N, hit, yield, P/L) con delta vs REAL
  3. Tabla comparativa por liga
  4. Tabla por camino (C1/C2/C2B/C3/C4)
  5. Detalle pick-a-pick de todos los simulados

Fecha: 2026-04-21.
"""
from collections import defaultdict
from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.comun.config_motor import get_param
from src.comun.reglas_actuales import evaluar_actual, evaluar_actual_ou
from src.persistencia.excel_estilos import fill


def _fecha_disp(fecha_raw):
    if not fecha_raw:
        return ""
    for fmt in ("%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(fecha_raw).strip(), fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return str(fecha_raw)[:10]


def _stake_kelly(prob, cuota, bankroll, fraccion_kelly, max_kelly_pct):
    """Replica motor_calculadora.calcular_stake_independiente con prob directo.
    Medio Kelly capado a max_kelly_pct."""
    if cuota <= 1 or prob <= 0:
        return 0.0
    kelly_full = (prob * cuota - 1) / (cuota - 1)
    if kelly_full <= 0:
        return 0.0
    fraccion = min(kelly_full * fraccion_kelly, max_kelly_pct)
    return round(bankroll * max(0, fraccion), 2)


def crear_hoja_resimulacion(wb, datos, bankroll):
    """Crea la hoja 'Si Hubiera' con resimulacion in-sample.

    datos: lista de filas tal como las devuelve _cargar_partidos del sincronizador.
           Columnas relevantes: 0=id, 1=fecha, 2=local, 3=visita, 4=pais,
           5-7=probs 1/x/2, 10=apuesta_1x2 (para comparar con real),
           14-16=cuotas 1/x/2, 19=estado, 20=gl, 21=gv.
    """
    ws = wb.create_sheet("Si Hubiera")

    FONT_TITLE = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    FONT_HDR   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    FONT_KPI   = Font(name='Arial', bold=True, size=10)
    FONT_D     = Font(name='Arial', size=10)
    FONT_SUB   = Font(name='Arial', italic=True, size=9, color='595959')

    FILL_TITULO   = fill('1F4E79')
    FILL_HDR_REAL = fill('808080')
    FILL_HDR_SIM  = fill('2E75B6')
    FILL_HDR_DLT  = fill('7030A0')
    FILL_HDR_CAM  = fill('548235')
    FILL_NEUTRO   = fill('F5F5F5')
    FILL_BLANCO   = fill('FFFFFF')
    FILL_GANADA   = fill('C6EFCE')
    FILL_PERDIDA  = fill('FFC7CE')
    FILL_AVISO    = fill('FFE699')

    BORDER = Border(
        left=Side(style='thin',  color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin',   color='D9D9D9'),
        bottom=Side(style='thin',color='D9D9D9'),
    )

    # Anchos (11 cols para detalle)
    widths = [12, 28, 12, 9, 8, 8, 10, 10, 10, 11, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # --- Parametros Kelly (leidos de configuracion) ---
    fraccion_kelly  = float(get_param('fraccion_kelly',      default=0.50) or 0.50)
    max_kelly_pct   = float(get_param('max_kelly_pct_normal', default=0.025) or 0.025)

    # --- Titulo ---
    ws.merge_cells('A1:K1')
    c = ws.cell(1, 1,
                f"RESIMULACION — Si siempre hubiera apostado con las reglas actuales (Fase 3.3.5) "
                f"| Bankroll base ${bankroll:,.0f} | Medio Kelly {fraccion_kelly:.0%} cap {max_kelly_pct:.1%}")
    c.font = FONT_TITLE; c.fill = FILL_TITULO
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # --- Advertencia ---
    ws.merge_cells('A2:K2')
    c = ws.cell(2, 1,
                "ADVERTENCIA: backtest IN-SAMPLE. Los criterios fueron calibrados sobre estos mismos "
                "datos — el hit/yield aqui sobreestima la ventaja real. Visibilidad solamente; no usar para decidir LIVE.")
    c.font = FONT_SUB; c.fill = FILL_AVISO
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 28

    # --- Procesamiento ---
    real_por_liga = defaultdict(lambda: {'n': 0, 'g': 0})
    sim_por_liga  = defaultdict(lambda: {'n': 0, 'g': 0,
                                         'stake_tot': 0.0, 'pl_tot': 0.0,
                                         'caminos': defaultdict(lambda: [0, 0])})
    picks_simulados = []

    # Fase 1: recolectar picks simulados + acumular REAL por liga
    candidatos = []  # picks que pasaron reglas (sin stake/pl aun, orden natural)
    for rd in datos:
        estado = rd[19]
        gl, gv = rd[20], rd[21]
        if estado != 'Liquidado' or gl is None or gv is None:
            continue
        pais  = rd[4]
        p1, px, p2 = rd[5], rd[6], rd[7]
        po, pu     = rd[8], rd[9]
        c1, cx, c2 = rd[14], rd[15], rd[16]
        co, cu     = rd[17], rd[18]
        ap_real    = str(rd[10] or '')
        ap_ou_real = str(rd[11] or '')
        fecha      = rd[1]
        # xg (cols 26/27) agregados al final de _cargar_partidos; pueden faltar en datos viejos
        xg_local  = rd[26] if len(rd) > 26 else None
        xg_visita = rd[27] if len(rd) > 27 else None

        tiene_1x2 = all(isinstance(x, (int, float)) and x > 0 for x in [p1, px, p2, c1, cx, c2])
        tiene_ou  = all(isinstance(x, (int, float)) and x > 0 for x in [po, pu, co, cu])

        # REAL 1X2
        if tiene_1x2 and ('[GANADA]' in ap_real or '[PERDIDA]' in ap_real):
            real_por_liga[pais]['n'] += 1
            if '[GANADA]' in ap_real:
                real_por_liga[pais]['g'] += 1
        # REAL O/U
        if '[GANADA]' in ap_ou_real or '[PERDIDA]' in ap_ou_real:
            real_por_liga[pais]['n'] += 1
            if '[GANADA]' in ap_ou_real:
                real_por_liga[pais]['g'] += 1

        # SIMULADO 1X2
        if tiene_1x2:
            pick, cuota, camino = evaluar_actual(p1, px, p2, c1, cx, c2, pais)
            if pick is not None:
                prob_pick = {'LOCAL': p1, 'EMPATE': px, 'VISITA': p2}[pick]
                gana = ((pick == 'LOCAL'  and gl >  gv) or
                        (pick == 'EMPATE' and gl == gv) or
                        (pick == 'VISITA' and gl <  gv))
                candidatos.append({
                    'fecha': fecha, 'local': rd[2], 'visita': rd[3], 'pais': pais,
                    'pick': pick, 'cuota': cuota, 'camino': camino, 'mercado': '1X2',
                    'gl': gl, 'gv': gv, 'gana': gana, 'prob': prob_pick,
                })

        # SIMULADO O/U 2.5
        if tiene_ou:
            pick_ou, cuota_ou, _ = evaluar_actual_ou(po, pu, co, cu, xg_local, xg_visita)
            if pick_ou is not None:
                prob_ou = po if pick_ou == 'OVER 2.5' else pu
                total_goles = gl + gv
                gana_ou = ((pick_ou == 'OVER 2.5'  and total_goles > 2.5) or
                           (pick_ou == 'UNDER 2.5' and total_goles < 2.5))
                candidatos.append({
                    'fecha': fecha, 'local': rd[2], 'visita': rd[3], 'pais': pais,
                    'pick': pick_ou, 'cuota': cuota_ou, 'camino': 'OU', 'mercado': 'O/U',
                    'gl': gl, 'gv': gv, 'gana': gana_ou, 'prob': prob_ou,
                })

    # Fase 2: ordenar cronologicamente y aplicar COMPOUNDING
    # bankroll_running arranca en bankroll base y se actualiza con cada P/L
    candidatos.sort(key=lambda x: str(x['fecha'] or ''))
    bankroll_running = float(bankroll)
    for p in candidatos:
        stake = _stake_kelly(p['prob'], p['cuota'], bankroll_running,
                             fraccion_kelly, max_kelly_pct)
        pl = round(stake * (p['cuota'] - 1), 2) if p['gana'] else round(-stake, 2)
        bankroll_running = round(bankroll_running + pl, 2)
        p['stake']  = stake
        p['pl']     = pl
        p['equity'] = bankroll_running

        pa = p['pais']
        sim_por_liga[pa]['n'] += 1
        sim_por_liga[pa]['stake_tot'] += stake
        sim_por_liga[pa]['pl_tot'] += pl
        sim_por_liga[pa]['caminos'][p['camino']][0] += 1
        if p['gana']:
            sim_por_liga[pa]['g'] += 1
            sim_por_liga[pa]['caminos'][p['camino']][1] += 1
        picks_simulados.append(p)

    # --- Totales globales ---
    total_r   = sum(v['n'] for v in real_por_liga.values())
    total_r_g = sum(v['g'] for v in real_por_liga.values())
    total_s   = sum(v['n'] for v in sim_por_liga.values())
    total_s_g = sum(v['g'] for v in sim_por_liga.values())
    total_s_stake = sum(v['stake_tot'] for v in sim_por_liga.values())
    total_s_pl    = sum(v['pl_tot']    for v in sim_por_liga.values())
    hit_r = total_r_g / total_r if total_r else 0
    hit_s = total_s_g / total_s if total_s else 0
    yld_s = total_s_pl / total_s_stake if total_s_stake else 0
    roi_bankroll = total_s_pl / bankroll if bankroll else 0

    # --- KPI globales (fila 4-5 headers, 6 datos) ---
    row = 4
    ws.merge_cells(f'A{row}:K{row}')
    c = ws.cell(row, 1, "RESUMEN GLOBAL")
    c.font = FONT_HDR; c.fill = FILL_TITULO
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 18
    row += 1

    kpi_hdr = ['KPI', 'REAL (historico)', 'SIMULADO (reglas actuales)', 'DELTA']
    kpi_widths = ['A', 'B:D', 'E:H', 'I:K']
    for ci_rng, txt, f in zip(kpi_widths, kpi_hdr,
                              [fill('4472C4'), FILL_HDR_REAL, FILL_HDR_SIM, FILL_HDR_DLT]):
        if ':' in ci_rng:
            ws.merge_cells(f'{ci_rng[0]}{row}:{ci_rng[2]}{row}')
            start = ord(ci_rng[0]) - 64
        else:
            start = ord(ci_rng) - 64
        c = ws.cell(row, start, txt)
        c.font = FONT_HDR; c.fill = f; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    def _kpi_row(label, real_val, sim_val, fmt='num', delta_fmt=None):
        ws.row_dimensions[row].height = 16
        c = ws.cell(row, 1, label)
        c.font = FONT_KPI; c.fill = FILL_NEUTRO; c.border = BORDER
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        # REAL (merge B:D)
        ws.merge_cells(f'B{row}:D{row}')
        cR = ws.cell(row, 2, real_val)
        cR.font = FONT_D; cR.border = BORDER
        cR.alignment = Alignment(horizontal='center', vertical='center')
        if fmt == 'pct': cR.number_format = '0.0%'
        elif fmt == 'cur': cR.number_format = '#,##0.00'
        elif fmt == 'int': cR.number_format = '0'

        # SIMULADO (merge E:H)
        ws.merge_cells(f'E{row}:H{row}')
        cS = ws.cell(row, 5, sim_val)
        cS.font = FONT_D; cS.border = BORDER
        cS.alignment = Alignment(horizontal='center', vertical='center')
        if fmt == 'pct': cS.number_format = '0.0%'
        elif fmt == 'cur': cS.number_format = '#,##0.00'
        elif fmt == 'int': cS.number_format = '0'

        # DELTA (merge I:K)
        ws.merge_cells(f'I{row}:K{row}')
        if isinstance(real_val, (int, float)) and isinstance(sim_val, (int, float)):
            delta = sim_val - real_val
            cD = ws.cell(row, 9, delta)
            cD.font = Font(name='Arial', bold=True, size=10,
                           color='006100' if delta > 0 else ('9C0006' if delta < 0 else '000000'))
            cD.fill = FILL_GANADA if delta > 0 else (FILL_PERDIDA if delta < 0 else FILL_NEUTRO)
            if delta_fmt: cD.number_format = delta_fmt
            elif fmt == 'pct': cD.number_format = '+0.0%;-0.0%;0.0%'
            elif fmt == 'cur': cD.number_format = '+#,##0.00;-#,##0.00'
            elif fmt == 'int': cD.number_format = '+0;-0;0'
        else:
            cD = ws.cell(row, 9, '—')
            cD.font = FONT_D; cD.fill = FILL_NEUTRO
        cD.border = BORDER
        cD.alignment = Alignment(horizontal='center', vertical='center')

    _kpi_row("N picks",        total_r,   total_s,  fmt='int'); row += 1
    _kpi_row("Ganados",        total_r_g, total_s_g, fmt='int'); row += 1
    _kpi_row("Hit rate",               hit_r, hit_s,           fmt='pct'); row += 1
    _kpi_row("Volumen apostado ($)",   0,     total_s_stake,   fmt='cur'); row += 1
    _kpi_row("Yield (P/L / volumen)",  0,     yld_s,           fmt='pct'); row += 1
    _kpi_row("P/L neto ($)",           0,     total_s_pl,      fmt='cur'); row += 1
    _kpi_row("ROI sobre bankroll",     0,     roi_bankroll,    fmt='pct'); row += 1
    _kpi_row("Equity final ($)",       bankroll, bankroll + total_s_pl, fmt='cur'); row += 1
    row += 1

    # --- Tabla por liga ---
    ws.merge_cells(f'A{row}:K{row}')
    c = ws.cell(row, 1, "DETALLE POR LIGA")
    c.font = FONT_HDR; c.fill = FILL_TITULO
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    hdrs = ['Liga', 'N real', 'Hit% real', 'N sim', 'G sim', 'Hit% sim',
            'Volumen $', 'P/L $', 'Yield sim', 'dN', 'dHit']
    fills_ = ([fill('4472C4')] + [FILL_HDR_REAL]*2 + [FILL_HDR_SIM]*5
              + [FILL_HDR_DLT]*2)
    # Expandir a 11 cols para acomodar volumen + P/L
    for ci, (h, f) in enumerate(zip(hdrs, fills_), 1):
        c = ws.cell(row, ci, h)
        c.font = FONT_HDR; c.fill = f; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    ligas_todas = sorted(set(list(real_por_liga.keys()) + list(sim_por_liga.keys())))
    for liga in ligas_todas:
        r_ = real_por_liga.get(liga, {'n': 0, 'g': 0})
        s_ = sim_por_liga.get(liga, {'n': 0, 'g': 0, 'stake_tot': 0.0, 'pl_tot': 0.0})
        hit_rl = r_['g'] / r_['n'] if r_['n'] else 0
        hit_sl = s_['g'] / s_['n'] if s_['n'] else 0
        yld_sl = s_['pl_tot'] / s_['stake_tot'] if s_['stake_tot'] else 0
        dN     = s_['n'] - r_['n']
        dHit   = hit_sl - hit_rl if (r_['n'] and s_['n']) else 0

        bg = FILL_BLANCO if row % 2 == 0 else FILL_NEUTRO
        vals = [
            (liga, None),                 (r_['n'], '0'),              (hit_rl, '0.0%'),
            (s_['n'], '0'),               (s_['g'], '0'),              (hit_sl, '0.0%'),
            (s_['stake_tot'], '#,##0.00'),(s_['pl_tot'], '+#,##0.00;-#,##0.00'),
            (yld_sl, '+0.0%;-0.0%;0.0%'), (dN, '+0;-0;0'),             (dHit, '+0.0%;-0.0%;0.0%'),
        ]
        for ci, (v, fmt) in enumerate(vals, 1):
            c = ws.cell(row, ci, v)
            c.font = FONT_D; c.fill = bg; c.border = BORDER
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                    vertical='center')
            if fmt: c.number_format = fmt
            if ci == 8 and isinstance(v, (int, float)):  # P/L
                c.fill = FILL_GANADA if v > 0 else (FILL_PERDIDA if v < 0 else bg)
            if ci == 9 and isinstance(v, (int, float)):  # Yield
                c.fill = FILL_GANADA if v > 0 else (FILL_PERDIDA if v < 0 else bg)
        ws.row_dimensions[row].height = 15
        row += 1
    row += 1

    # --- Tabla por camino ---
    ws.merge_cells(f'A{row}:K{row}')
    c = ws.cell(row, 1, "PICKS SIMULADOS POR CAMINO (formato: n_picks/ganados)")
    c.font = FONT_HDR; c.fill = FILL_HDR_CAM
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    cam_hdrs = ['Liga', 'C1', 'C2', 'C2B', 'C3', 'C4', 'OU', 'Total']
    for ci, h in enumerate(cam_hdrs, 1):
        c = ws.cell(row, ci, h)
        c.font = FONT_HDR; c.fill = FILL_HDR_CAM; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    for liga in ligas_todas:
        s_ = sim_por_liga.get(liga, {'caminos': {}})
        cams = s_.get('caminos', {})
        c1 = cams.get('C1', [0, 0])
        c2 = cams.get('C2', [0, 0])
        c2b = cams.get('C2B', [0, 0])
        c3 = cams.get('C3', [0, 0])
        c4 = cams.get('C4', [0, 0])
        ou = cams.get('OU', [0, 0])
        total = c1[0] + c2[0] + c2b[0] + c3[0] + c4[0] + ou[0]
        if total == 0:
            continue
        bg = FILL_BLANCO if row % 2 == 0 else FILL_NEUTRO

        def _fmt_cam(x):
            return f'{x[0]}/{x[1]}' if x[0] else '-'

        vals = [liga, _fmt_cam(c1), _fmt_cam(c2), _fmt_cam(c2b),
                _fmt_cam(c3), _fmt_cam(c4), _fmt_cam(ou), total]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row, ci, v)
            c.font = FONT_D; c.fill = bg; c.border = BORDER
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                    vertical='center')
        ws.row_dimensions[row].height = 15
        row += 1
    row += 1

    # --- Detalle de cada pick simulado ---
    ws.merge_cells(f'A{row}:K{row}')
    c = ws.cell(row, 1, f"DETALLE DE PICKS SIMULADOS (n={len(picks_simulados)})")
    c.font = FONT_HDR; c.fill = FILL_TITULO
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    det_hdrs = ['Fecha', 'Partido', 'Liga', 'Pick', 'Cuota', 'Camino',
                'Goles', 'Resultado', 'Stake $', 'P/L $', 'Equity $']
    for ci, h in enumerate(det_hdrs, 1):
        c = ws.cell(row, ci, h)
        c.font = FONT_HDR; c.fill = FILL_HDR_SIM; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 1

    for p in picks_simulados:  # ya vienen en orden cronologico
        bg = FILL_BLANCO if row % 2 == 0 else FILL_NEUTRO
        res_fill = FILL_GANADA if p['gana'] else FILL_PERDIDA
        pl_fill  = FILL_GANADA if p['pl'] > 0 else FILL_PERDIDA

        vals = [
            (_fecha_disp(p['fecha']), None, bg),
            (f"{p['local']} vs {p['visita']}", None, bg),
            (p['pais'], None, bg),
            (p['pick'], None, bg),
            (p['cuota'], '0.00', bg),
            (p['camino'], None, bg),
            (f"{p['gl']}-{p['gv']}", None, bg),
            ('GANADA' if p['gana'] else 'PERDIDA', None, res_fill),
            (p['stake'],  '#,##0.00', bg),
            (p['pl'],     '+#,##0.00;-#,##0.00', pl_fill),
            (p['equity'], '#,##0.00', bg),
        ]
        for ci, (v, fmt, cfill) in enumerate(vals, 1):
            c = ws.cell(row, ci, v)
            c.font = FONT_D; c.fill = cfill; c.border = BORDER
            c.alignment = Alignment(
                horizontal='left' if ci == 2 else 'center', vertical='center')
            if fmt: c.number_format = fmt
        ws.row_dimensions[row].height = 14
        row += 1

    # Pie
    row += 1
    ws.merge_cells(f'A{row}:K{row}')
    c = ws.cell(row, 1,
                f"Stakes calculados con Medio Kelly ({fraccion_kelly:.0%}) capado a {max_kelly_pct:.1%} del bankroll, "
                f"CON compounding: cada pick se apuesta sobre bankroll_running = bankroll_base + P/L acumulado al momento del pick. "
                f"Equity arranca en ${bankroll:,.0f}. No se aplica ajuste de covarianza (multiples picks/dia).")
    c.font = FONT_SUB
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 32
