"""
Hoja "LIVE" — solo partidos futuros con stake > 0 (apuestas reales activas).

Dividida por liga para reducir ruido visual. Cada liga trae:
  - Header con flag + nombre + N picks + stake total
  - Tabla de picks: fecha, partido, mercado, pick, cuota, probs, EV%, stake, P/L potencial
  - Fill pastel del pais (paleta PAISES_CF)

Vista de trabajo operativa del dia: qué tengo apostado y cuánto me juego.
Se excluyen picks de pretest (stake=0) y liquidados (ya cerrados).

Fecha: 2026-04-24.
"""
from collections import defaultdict
from datetime import datetime

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from src.persistencia.excel_estilos import (
    FONT_HEADER, FONT_DATA, FILL_HEADER, PAISES_CF,
    FILL_APOSTAR, FILL_VERDE, FILL_AMARILLO, FILL_ROJO,
    ALIGN_CENTER, ALIGN_LEFT, BORDER_THIN, fill,
)


FLAGS = {
    "Argentina": "AR", "Brasil": "BR", "Uruguay": "UY", "Chile": "CL",
    "Peru": "PE", "Ecuador": "EC", "Colombia": "CO", "Bolivia": "BO",
    "Venezuela": "VE", "Espana": "ES", "Italia": "IT", "Alemania": "DE",
    "Francia": "FR", "Inglaterra": "EN", "Turquia": "TR", "Noruega": "NO",
}

# EV bands (coherente con dashboard Streamlit)
FILL_EV_EXCELENTE = fill('0A5C2A')   # verde oscuro >= 15%
FILL_EV_BUENO     = FILL_VERDE       #             >=  8%
FILL_EV_OK        = FILL_AMARILLO    #             >=  3%
FILL_EV_FLOJO     = FILL_ROJO        #             <   3%

FONT_TITULO  = Font(name='Arial', bold=True, size=16, color='1F4E79')
FONT_LIGA    = Font(name='Arial', bold=True, size=12, color='FFFFFF')
FONT_NEGRITA = Font(name='Arial', bold=True, size=10)
FILL_TITULO  = fill('1F4E79')

COL_WIDTHS = [14, 32, 9, 14, 8, 10, 10, 9, 12, 13]


def _fecha_display(fecha_raw):
    if not fecha_raw:
        return ""
    s = str(fecha_raw).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m %H:%M")
        except ValueError:
            continue
    return s[:16]


def _parse_pick_1x2(apuesta_str, cuota_1, cuota_x, cuota_2, prob_1, prob_x, prob_2):
    """Devuelve (pick_label, cuota, prob_modelo) o None si no aplica."""
    if not apuesta_str or not apuesta_str.startswith('[APOSTAR]'):
        return None
    s = apuesta_str.upper()
    if 'LOCAL' in s:
        return ('LOCAL', cuota_1, prob_1)
    if 'EMPATE' in s:
        return ('EMPATE', cuota_x, prob_x)
    if 'VISITA' in s:
        return ('VISITA', cuota_2, prob_2)
    return None


def _parse_pick_ou(apuesta_str, cuota_o25, cuota_u25, prob_o25, prob_u25):
    if not apuesta_str or not apuesta_str.startswith('[APOSTAR]'):
        return None
    s = apuesta_str.upper()
    if 'OVER' in s:
        return ('OVER 2.5', cuota_o25, prob_o25)
    if 'UNDER' in s:
        return ('UNDER 2.5', cuota_u25, prob_u25)
    return None


def _fill_ev(ev_pct):
    if ev_pct is None:
        return None
    if ev_pct >= 15: return FILL_EV_EXCELENTE
    if ev_pct >= 8:  return FILL_EV_BUENO
    if ev_pct >= 3:  return FILL_EV_OK
    return FILL_EV_FLOJO


def _build_rows(datos):
    """Extrae picks vivos (stake>0, no liquidados) desde la lista cruda de sincronizador.

    `datos` viene con este orden (ver motor_sincronizador._cargar_partidos):
      0=id, 1=fecha, 2=local, 3=visita, 4=pais,
      5=prob_1, 6=prob_x, 7=prob_2, 8=prob_o25, 9=prob_u25,
      10=apuesta_1x2, 11=apuesta_ou, 12=stake_1x2, 13=stake_ou,
      14=cuota_1, 15=cuota_x, 16=cuota_2, 17=cuota_o25, 18=cuota_u25,
      19=estado, ...
    """
    rows_por_liga = defaultdict(list)
    for r in datos:
        estado = r[19]
        if estado == 'Liquidado':
            continue
        local, visita, pais = r[2], r[3], r[4]
        fecha = r[1]
        apuesta_1x2 = r[10]
        apuesta_ou  = r[11]
        stake_1x2   = r[12] or 0
        stake_ou    = r[13] or 0
        if stake_1x2 <= 0 and stake_ou <= 0:
            continue

        if stake_1x2 > 0:
            parsed = _parse_pick_1x2(apuesta_1x2, r[14], r[15], r[16], r[5], r[6], r[7])
            if parsed:
                pick, cuota, prob = parsed
                rows_por_liga[pais].append({
                    'fecha': fecha, 'partido': f"{local} vs {visita}",
                    'mercado': '1X2', 'pick': pick, 'cuota': cuota,
                    'prob_modelo': prob, 'stake': stake_1x2,
                })
        if stake_ou > 0:
            parsed = _parse_pick_ou(apuesta_ou, r[17], r[18], r[8], r[9])
            if parsed:
                pick, cuota, prob = parsed
                rows_por_liga[pais].append({
                    'fecha': fecha, 'partido': f"{local} vs {visita}",
                    'mercado': 'O/U', 'pick': pick, 'cuota': cuota,
                    'prob_modelo': prob, 'stake': stake_ou,
                })
    return rows_por_liga


def crear_hoja_live(wb, datos, apuestas_live=None):
    """Crea la hoja 'LIVE' dividida por liga con picks vivos (stake>0).

    datos: filas tal como las devuelve _cargar_partidos del sincronizador.
    apuestas_live: dict pais->bool con el estado LIVE/pretest por pais (para tag visual).
    """
    apuestas_live = apuestas_live or {}
    ws = wb.create_sheet("LIVE", 0)  # insertar primero para que sea la primera pestana
    # Nota: se insertara al frente al final para no interferir con crear_hoja_dashboard

    rows_por_liga = _build_rows(datos)

    # Ajustar anchos de columna
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Titulo ---
    ws.cell(row=1, column=1, value="APUESTAS LIVE").font = FONT_TITULO
    ws.cell(row=2, column=1, value=f"Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = FONT_DATA

    total_picks = sum(len(v) for v in rows_por_liga.values())
    total_stake = sum(p['stake'] for picks in rows_por_liga.values() for p in picks)
    ws.cell(row=3, column=1,
            value=f"{total_picks} picks vivos · ${total_stake:,.0f} en stake total").font = FONT_NEGRITA

    row = 5
    if not rows_por_liga:
        ws.cell(row=row, column=1, value="No hay apuestas vivas (todas las ligas en pretest)").font = FONT_DATA
        return ws

    # Map pais -> fill PAISES_CF (color pastel)
    pais_fill = dict(PAISES_CF)

    # Ordenar ligas: primero las que aparecen en apuestas_live (LIVE 1X2), luego por cantidad de picks
    ligas_ordenadas = sorted(
        rows_por_liga.keys(),
        key=lambda p: (not apuestas_live.get(p, False), -len(rows_por_liga[p])),
    )

    for pais in ligas_ordenadas:
        picks = rows_por_liga[pais]
        if not picks:
            continue

        # --- Header de liga ---
        flag = FLAGS.get(pais, '--')
        live_tag = " · LIVE" if apuestas_live.get(pais, False) else ""
        stake_liga = sum(p['stake'] for p in picks)
        titulo_liga = f"  {flag}  {pais}  —  {len(picks)} picks · ${stake_liga:,.0f}{live_tag}"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_WIDTHS))
        c = ws.cell(row=row, column=1, value=titulo_liga)
        c.font = FONT_LIGA
        c.fill = FILL_TITULO
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        # --- Header de tabla ---
        headers = [
            'Fecha', 'Partido', 'Mercado', 'Pick', 'Cuota',
            'Prob modelo', 'Prob mercado', 'EV %', 'Stake $', 'P/L potencial',
        ]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = BORDER_THIN
        row += 1

        # --- Rows ---
        liga_fill = pais_fill.get(pais)
        picks_sorted = sorted(picks, key=lambda p: (p['fecha'] or '', p['mercado']))
        for p in picks_sorted:
            cuota = float(p['cuota'] or 0)
            prob_modelo = float(p['prob_modelo'] or 0)
            prob_mercado = (1.0 / cuota) if cuota > 0 else 0
            ev_pct = (prob_modelo * cuota - 1.0) * 100.0 if cuota > 0 else 0
            pl_pot = p['stake'] * (cuota - 1.0) if cuota > 0 else 0

            values = [
                _fecha_display(p['fecha']),
                p['partido'],
                p['mercado'],
                p['pick'],
                cuota,
                prob_modelo,
                prob_mercado,
                ev_pct,
                p['stake'],
                pl_pot,
            ]
            formats = [None, None, None, None, '0.00', '0.0%', '0.0%',
                       '+0.0"%";-0.0"%";0.0"%"', '$#,##0', '$#,##0']
            for col_idx, (val, fmt) in enumerate(zip(values, formats), start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = FONT_DATA
                c.border = BORDER_THIN
                c.alignment = ALIGN_LEFT if col_idx in (2, 4) else ALIGN_CENTER
                if fmt:
                    c.number_format = fmt
                if liga_fill is not None:
                    c.fill = liga_fill

            # Sobrescribir fill de EV con su semáforo
            ev_cell = ws.cell(row=row, column=8)
            ev_fill = _fill_ev(ev_pct)
            if ev_fill is not None:
                ev_cell.fill = ev_fill
            ev_cell.font = Font(name='Arial', size=10, bold=True,
                                color='FFFFFF' if ev_pct >= 8 else '000000')

            # Stake en negrita
            ws.cell(row=row, column=9).font = FONT_NEGRITA

            row += 1

        row += 1  # fila espaciadora entre ligas

    # Freeze panes bajo el titulo/KPIs (arriba siempre visibles)
    ws.freeze_panes = 'A5'
    return ws
