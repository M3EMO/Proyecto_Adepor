"""
Hoja "Dashboard" — la que abre primero con KPIs financieros y estadisticos.

Secciones:
  - Resultados financieros (P/L, yield, volumen, N)
  - Tasa de acierto (sistema, apuestas, all)
  - Estadistica inferencial (T-score, p-value, Kelly)
  - Calibracion (Brier sistema/casa/global)
  - Estrategia activa (resumen de filtros)

Extraido del motor_sincronizador.py monolitico en fase 4 (2026-04-21).
"""
from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side

from src.persistencia.excel_estilos import fill, FILL_NEUTRO
from src.persistencia.excel_metricas import semaforo


def crear_hoja_dashboard(wb, metricas, bankroll, apuestas_live_por_liga=None):
    """Crea la hoja Dashboard como primera pestana del workbook.

    apuestas_live_por_liga: dict pais -> bool (True = LIVE, False = pretest).
    Si None, se omite la columna 'estado' de la tabla por liga.
    """
    apuestas_live_por_liga = apuestas_live_por_liga or {}
    ws = wb.create_sheet("Dashboard", 0)

    FONT_TITLE = Font(name='Arial', bold=True, color='FFFFFF', size=13)
    FONT_SEC   = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    FONT_KPI   = Font(name='Arial', bold=True, size=10)
    FONT_VAL   = Font(name='Arial', size=10)
    FONT_SUB   = Font(name='Arial', italic=True, size=9, color='595959')
    FILL_TITLE_D = fill('1F4E79')
    FILL_SEC_D   = fill('4472C4')
    FILL_HDR_COL = fill('2E75B6')
    BORDER_DB = Border(
        left=Side(style='thin', color='9DC3E6'),
        right=Side(style='thin', color='9DC3E6'),
        top=Side(style='thin',  color='9DC3E6'),
        bottom=Side(style='thin', color='9DC3E6'),
    )

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # --- Titulo ---
    ws.merge_cells('A1:D1')
    c = ws.cell(1, 1, "DASHBOARD DE RENDIMIENTO")
    c.font = FONT_TITLE; c.fill = FILL_TITLE_D
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    c = ws.cell(2, 1,
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
                f"Bankroll: ${bankroll:,.2f}")
    c.font = FONT_SUB; c.fill = fill('D6E4F0')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 15

    # --- Headers columnas fila 3 ---
    for ci, h in enumerate(['Metrica', 'Total', '1X2', 'O/U'], 1):
        c = ws.cell(3, ci, h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = FILL_HDR_COL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER_DB
    ws.row_dimensions[3].height = 20

    m = metricas
    t_all, t_1x2, t_ou = m['total'], m['1x2'], m['ou']
    row = [4]  # mutable para closures

    def _sep(titulo):
        ws.merge_cells(f'A{row[0]}:D{row[0]}')
        c = ws.cell(row[0], 1, titulo)
        c.font = FONT_SEC; c.fill = FILL_SEC_D
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = BORDER_DB
        ws.row_dimensions[row[0]].height = 14
        row[0] += 1

    def _fila(metrica, vals, fmts, fills):
        r = row[0]
        bg = fill('EBF3FB') if r % 2 == 0 else fill('FFFFFF')
        ws.row_dimensions[r].height = 18
        c = ws.cell(r, 1, metrica); c.font = FONT_KPI; c.fill = bg; c.border = BORDER_DB
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        for ci, (val, fmt, fill_) in enumerate(zip(vals, fmts, fills), 2):
            cell = ws.cell(r, ci, val); cell.font = FONT_VAL; cell.border = BORDER_DB
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill_ if fill_ else bg
            if   fmt == 'pct': cell.number_format = '0.00%'
            elif fmt == 'cur': cell.number_format = '#,##0.00'
            elif fmt == 'd4':  cell.number_format = '0.0000'
            elif fmt == 'd2':  cell.number_format = '0.00'
            elif fmt == 'int': cell.number_format = '0'
        row[0] += 1

    NA = '—'

    # ==========================================================================
    # RESULTADOS FINANCIEROS
    # ==========================================================================
    _sep("  RESULTADOS FINANCIEROS")

    _fila("Ganancia neta",
          (t_all['pl'], t_1x2['pl'], t_ou['pl']),
          ('cur', 'cur', 'cur'),
          (semaforo(t_all['pl'], 0, 0),
           semaforo(t_1x2['pl'], 0, 0),
           semaforo(t_ou['pl'],  0, 0)))

    _fila("Yield",
          (t_all['yield'], t_1x2['yield'], t_ou['yield']),
          ('pct', 'pct', 'pct'),
          (semaforo(t_all['yield'], 0.05, 0),
           semaforo(t_1x2['yield'], 0.05, 0),
           semaforo(t_ou['yield'],  0.05, 0)))

    _fila("Volumen apostado",
          (t_all['vol'], t_1x2['vol'], t_ou['vol']),
          ('cur', 'cur', 'cur'),
          (FILL_NEUTRO, FILL_NEUTRO, FILL_NEUTRO))

    _fila("N apuestas liquidadas",
          (t_all['n'], t_1x2['n'], t_ou['n']),
          ('int', 'int', 'int'),
          (FILL_NEUTRO, FILL_NEUTRO, FILL_NEUTRO))

    # ==========================================================================
    # TASA DE ACIERTO
    # ==========================================================================
    _sep("  TASA DE ACIERTO")

    ap = m['acierto_partidos']
    pt = m['pred_total']
    _fila(f"% Acierto P  (col. Acierto, {m['pred_aciertos']}/{pt} partidos)",
          (ap, ap, '—'),
          ('pct', 'pct', ''),
          (semaforo(ap, 0.55, 0.45),
           semaforo(ap, 0.55, 0.45),
           FILL_NEUTRO))

    _fila("% Acierto $  (apuestas ganadoras / total apostado)",
          (t_all['acierto_bets'], t_1x2['acierto_bets'], t_ou['acierto_bets']),
          ('pct', 'pct', 'pct'),
          (semaforo(t_all['acierto_bets'], 0.55, 0.45),
           semaforo(t_1x2['acierto_bets'], 0.55, 0.45),
           semaforo(t_ou['acierto_bets'],  0.55, 0.45)))

    # % Acierto all: N/A en pretest (t['n']==0 = sin apuestas con stake real).
    ap_sistema = m['acierto_partidos']
    pretest_total = (t_all['n'] == 0)
    pretest_1x2   = (t_1x2['n'] == 0)
    pretest_ou    = (t_ou['n']  == 0)
    NA_PRE = 'N/A (pretest)'
    all_total = NA_PRE if pretest_total else (ap_sistema + t_all['acierto_bets']) / 2
    all_1x2   = NA_PRE if pretest_1x2   else (ap_sistema + t_1x2['acierto_bets']) / 2
    all_ou    = NA_PRE if pretest_ou    else t_ou['acierto_bets']
    _fila("% Acierto all  (promedio sistema + apuestas; N/A hasta que corra dinero)",
          (all_total, all_1x2, all_ou),
          ('' if pretest_total else 'pct',
           '' if pretest_1x2   else 'pct',
           '' if pretest_ou    else 'pct'),
          (FILL_NEUTRO if pretest_total else semaforo(all_total, 0.55, 0.45),
           FILL_NEUTRO if pretest_1x2   else semaforo(all_1x2,   0.55, 0.45),
           FILL_NEUTRO if pretest_ou    else semaforo(all_ou,    0.55, 0.45)))

    # ==========================================================================
    # ESTADISTICA INFERENCIAL
    # ==========================================================================
    _sep("  ESTADISTICA INFERENCIAL")

    _fila("T-score",
          (t_all['t'], t_1x2['t'], t_ou['t']),
          ('d2', 'd2', 'd2'),
          (semaforo(abs(t_all['t']), 2.0, 1.0),
           semaforo(abs(t_1x2['t']), 2.0, 1.0),
           semaforo(abs(t_ou['t']),  2.0, 1.0)))

    _fila("P-Value  (two-tailed, <0.05 = significativo)",
          (t_all['p'], t_1x2['p'], t_ou['p']),
          ('d4', 'd4', 'd4'),
          (semaforo(t_all['p'], 0.05, 0.10, mayor_es_mejor=False),
           semaforo(t_1x2['p'], 0.05, 0.10, mayor_es_mejor=False),
           semaforo(t_ou['p'],  0.05, 0.10, mayor_es_mejor=False)))

    fk = m['fraccion_kelly']
    _fila("Fraccion Kelly",
          (fk, fk, NA), ('pct', 'pct', ''),
          (FILL_NEUTRO, FILL_NEUTRO, FILL_NEUTRO))

    # ==========================================================================
    # CALIBRACION (Brier rango 0-2; aleatorio puro ~= 0.667)
    # ==========================================================================
    _sep("  CALIBRACION DEL MODELO  (Brier Score — rango 0 a 2, aleatorio ≈ 0.667, menor es mejor)")

    bs_s = m['bs_sis']; bs_c = m['bs_casa']; bs_g = m['bs_glob']

    _fila("BS Sistema  (promedio por partido, Dixon-Coles)",
          (bs_s, bs_s, NA), ('d4', 'd4', ''),
          (semaforo(bs_s, 0.50, 0.65, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    _fila("BS Casa  (promedio por partido, cuotas mercado)",
          (bs_c, bs_c, NA), ('d4', 'd4', ''),
          (semaforo(bs_c, 0.50, 0.65, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    _fila("BS Global  (Sistema - Casa, negativo = modelo supera mercado)",
          (bs_g, bs_g, NA), ('d4', 'd4', ''),
          (semaforo(bs_g, -0.02, 0.02, mayor_es_mejor=False), FILL_NEUTRO, FILL_NEUTRO))

    # Leyenda
    ws.merge_cells(f'A{row[0] + 1}:D{row[0] + 1}')
    c = ws.cell(row[0] + 1, 1,
                "Yield >5%=verde | 0-5%=amarillo | <0%=rojo    "
                "BS <0.50=verde | 0.50-0.65=amarillo | >0.65=rojo    "
                "P-Value <0.05=verde | 0.05-0.10=amarillo | >0.10=rojo    "
                "BS Global negativo = modelo supera al mercado")
    c.font = Font(name='Arial', italic=True, size=8, color='595959')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row[0] += 3

    # ==========================================================================
    # PERFORMANCE POR LIGA (N, hit%, yield, apostado, estado LIVE/pretest)
    # ==========================================================================
    por_liga = m.get('por_liga', {}) or {}
    if por_liga:
        ws.merge_cells(f'A{row[0]}:D{row[0]}')
        c = ws.cell(row[0], 1, "  PERFORMANCE POR LIGA  (apuestas liquidadas con stake real)")
        c.font = FONT_SEC; c.fill = fill('7F6000')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = BORDER_DB
        ws.row_dimensions[row[0]].height = 14
        row[0] += 1

        # Sub-header
        hdrs = ['Liga', 'N', 'Hit%', 'Yield%  /  Apostado  /  Estado']
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row[0], ci, h)
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            c.fill = fill('7F6000')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = BORDER_DB
        ws.row_dimensions[row[0]].height = 14
        row[0] += 1

        for liga, s in sorted(por_liga.items()):
            r = row[0]
            bg = fill('FFF9E6') if r % 2 == 0 else fill('FFFFFF')
            ws.row_dimensions[r].height = 16
            hit = s['g'] / s['n'] if s['n'] > 0 else 0
            yld = s['pl'] / s['vol'] if s['vol'] > 0 else 0
            live = apuestas_live_por_liga.get(liga, False)
            estado = 'LIVE' if live else 'pretest'
            estado_fill = fill('C6EFCE') if live else fill('FFE699')

            # col A: liga
            c = ws.cell(r, 1, liga); c.font = FONT_KPI; c.fill = bg; c.border = BORDER_DB
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            # col B: N
            c = ws.cell(r, 2, s['n']); c.font = FONT_VAL; c.fill = bg; c.border = BORDER_DB
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.number_format = '0'
            # col C: hit
            c = ws.cell(r, 3, hit); c.font = FONT_VAL; c.border = BORDER_DB
            c.fill = semaforo(hit, 0.55, 0.45) if s['n'] > 0 else FILL_NEUTRO
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.number_format = '0.0%'
            # col D: compuesto yield + apostado + estado
            info = f"{100*yld:+.1f}%  |  ${s['vol']:,.0f}  |  {estado}"
            c = ws.cell(r, 4, info); c.font = FONT_VAL; c.border = BORDER_DB
            c.fill = estado_fill if estado == 'LIVE' else (semaforo(yld, 0.05, 0) if s['n'] > 0 else bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            row[0] += 1

        # Leyenda
        ws.merge_cells(f'A{row[0]}:D{row[0]}')
        c = ws.cell(row[0], 1,
                    "LIVE = stake real activo (pretest gate superado). "
                    "pretest = picks generados pero stake=0 hasta N>=15 + hit>=55% + p<=0.30.")
        c.font = Font(name='Arial', italic=True, size=8, color='595959')
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        row[0] += 2

    # ==========================================================================
    # ESTRATEGIA ACTIVA (texto estatico con filtros)
    # ==========================================================================
    ws.merge_cells(f'A{row[0]}:D{row[0]}')
    c = ws.cell(row[0], 1, "  ESTRATEGIA ACTIVA  (Motor Calculadora V4.3)")
    c.font = FONT_SEC; c.fill = fill('375623')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = BORDER_DB
    ws.row_dimensions[row[0]].height = 14
    row[0] += 1

    FILL_CFG = fill('EBF3EB')
    FONT_CFG_K = Font(name='Arial', bold=True, size=9)
    FONT_CFG_V = Font(name='Arial', size=9)

    def _cfg(param, valor, nota=''):
        r = row[0]
        bg = FILL_CFG if r % 2 == 0 else fill('FFFFFF')
        ws.row_dimensions[r].height = 15
        ws.merge_cells(f'A{r}:B{r}')
        ck = ws.cell(r, 1, param)
        ck.font = FONT_CFG_K; ck.fill = bg; ck.border = BORDER_DB
        ck.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ws.merge_cells(f'C{r}:D{r}')
        cv = ws.cell(r, 3, valor)
        cv.font = FONT_CFG_V
        cv.fill = fill('C6EFCE') if nota == 'activo' else bg
        cv.border = BORDER_DB
        cv.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        row[0] += 1

    _cfg("Floor prob mínima",          "33%  — ningún outcome por debajo de este piso")
    _cfg("EV mínimo escalado",          "prob≥50%->3%  |  prob 40-50%->8%  |  prob 33-40%->12%")
    _cfg("Bloqueo empates",             "ACTIVO — sobreestimación sistémica +7.9% vs real", 'activo')
    _cfg("Camino 2B — Desacuerdo",      "modelo≠mercado + prob≥40% + div 15-30% + EV escalado")
    _cfg("Camino 3 — Alta Convicción",  "prob≥33% + EV≥100% + cuota≤8.0")
    _cfg("xG Margen O/U",               "apostar O/U solo si |xG_total − 2.5| ≥ 0.4 goles")
    _cfg("Margen predictivo 1X2",       "diferencia entre 1º y 2º prob del modelo ≥ 3%")
    _cfg("Divergencia normal",          "prob_modelo − prob_implícita_mercado ≤ 15%")
    _cfg("Techo cuota normal",          "≤ 5.0  (relajado a 8.0 en Caminos 2B y 3)")
    _cfg("Kelly fraccionado",           f"{m['fraccion_kelly']:.0%} del Kelly óptimo (Thorp 2006)")

    ws.freeze_panes = 'A4'
